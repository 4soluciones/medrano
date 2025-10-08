"""
Vistas para exportación de reportes a Excel
"""
import os
import decimal
import openpyxl
from decimal import Decimal
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q
from http import HTTPStatus

from medrano import settings
from .models import CashFlow
from ..sales.models import Order
from ..hrm.models import Subsidiary


@csrf_exempt
def export_sales_report_excel(request):
    """Exportar reporte de ventas a Excel"""
    if request.method == 'POST':
        try:
            # Obtener datos del reporte
            report_date = request.POST.get('report_date')
            subsidiary_id = request.POST.get('subsidiary')
            cash_id = request.POST.get('cash_account')
            subsidiary_obj = None
            
            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por sucursal
            if subsidiary_id and subsidiary_id != '0':
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    cash__subsidiary_id=subsidiary_id
                )
            else:
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']
            ).order_by('id')
            
            if subsidiary_id and subsidiary_id != '0':
                orders_of_day = orders_of_day.filter(subsidiary_id=subsidiary_id)
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # Crear estructura de datos para adelantos (ingresos del día)
            advances_grouped = {}
            for order in orders_of_day:
                order_cashflows_day = cashflows.filter(
                    order=order,
                    type='E',
                    transaction_date=report_date
                ).order_by('id')
                
                if order_cashflows_day.exists():
                    total_paid = sum(decimal.Decimal(cf.total) for cf in order_cashflows_day)
                    saldo = decimal.Decimal(order.total) - total_paid
                    is_paid_in_full = abs(saldo) < 0.01
                    
                    advances_grouped[order.id] = {
                        'order': order,
                        'cashflows': list(order_cashflows_day),
                        'total_advances': total_paid,
                        'saldo': saldo,
                        'is_paid_in_full': is_paid_in_full,
                        'cashflow_count': order_cashflows_day.count()
                    }
            
            # Preparar datos de saldos (cancelaciones)
            payments_cashflows = order_cashflows.filter(
                type='E',
                order_type_entry='T'
            ).exclude(
                order__register_date__gte=report_date
            )
            
            # Preparar datos de egresos
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'
            )
            
            # Calcular totales
            total_advances = sum(data['total_advances'] for data in advances_grouped.values())
            total_payments = payments_cashflows.aggregate(total=Sum('total'))['total'] or 0
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular totales por tipo de pago
            advances_efectivo = 0
            advances_yape = 0
            advances_deposito = 0
            
            for data in advances_grouped.values():
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        advances_efectivo += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        advances_yape += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'D':
                        advances_deposito += decimal.Decimal(cashflow.total)
            
            payments_efectivo = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
            payments_yape = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
            
            total_efectivo = advances_efectivo + payments_efectivo
            total_yape = advances_yape + payments_yape
            total_general = total_efectivo + total_yape
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Reporte {report_date}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill_primary = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
            header_fill_success = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            header_fill_danger = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            title_font = Font(bold=True, size=16, color="007bff")
            border = Border(
                left=Side(style='medium', color='adb5bd'),
                right=Side(style='medium', color='adb5bd'),
                top=Side(style='medium', color='adb5bd'),
                bottom=Side(style='medium', color='adb5bd')
            )
            
            # Título principal
            ws.merge_cells('A1:J1')
            ws['A1'] = f"TIENDA: {subsidiary_obj.name.upper() if subsidiary_obj else 'TODAS'} - DÍA: {datetime.strptime(report_date, '%Y-%m-%d').strftime('%d-%m-%Y')}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Sección de INGRESOS DEL DÍA
            ws['A3'] = "INGRESOS DEL DÍA"
            ws['A3'].font = header_font
            ws['A3'].fill = header_fill_primary
            ws.merge_cells('A3:J3')
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # Encabezados de ingresos
            income_headers = ['N° CPTE.', 'CLIENTE O RAZON SOCIAL', 'CANT.', 'DESCRIPCIÓN DEL PRODUCTO', 'USUARIO', 'TIPO PAGO', 'A CUENTA S/.', 'SALDO S/.', 'TOTAL S/.']
            for col, header in enumerate(income_headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_primary
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de ingresos del día
            row = 5
            for order_id, data in advances_grouped.items():
                if not data['is_paid_in_full']:  # Solo adelantos
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                            ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                            ws.cell(row=row, column=3, value=1).border = border  # Cantidad
                            # Descripción del producto
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            ws.cell(row=row, column=4, value=product_desc).border = border
                        else:
                            # Filas adicionales sin datos de orden
                            ws.cell(row=row, column=1, value="").border = border
                            ws.cell(row=row, column=2, value="").border = border
                            ws.cell(row=row, column=3, value="").border = border
                            ws.cell(row=row, column=4, value="").border = border
                        
                        # Datos del cashflow
                        ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                        
                        # Tipo de pago
                        payment_type = ""
                        if cashflow.way_to_pay == 'E':
                            payment_type = "EFECTIVO"
                        elif cashflow.way_to_pay == 'Y':
                            payment_type = "YAPE"
                        elif cashflow.way_to_pay == 'D':
                            payment_type = "DEPÓSITO"
                        ws.cell(row=row, column=6, value=payment_type).border = border
                        ws.cell(row=row, column=7, value=Decimal(cashflow.total)).border = border
                        
                        if i == 0:  # Solo en la primera fila
                            ws.cell(row=row, column=8, value=Decimal(data['saldo'])).border = border
                            ws.cell(row=row, column=9, value=Decimal(data['order'].total)).border = border
                        else:
                            ws.cell(row=row, column=8, value="").border = border
                            ws.cell(row=row, column=9, value="").border = border
                        
                row += 1
            
                # Pagos completos
                if data['is_paid_in_full']:
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                            ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                            ws.cell(row=row, column=3, value=1).border = border
                            # Descripción del producto
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            ws.cell(row=row, column=4, value=product_desc).border = border
                        else:
                            ws.cell(row=row, column=1, value="").border = border
                            ws.cell(row=row, column=2, value="").border = border
                            ws.cell(row=row, column=3, value="").border = border
                            ws.cell(row=row, column=4, value="").border = border
                        
                        ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                        
                        payment_type = ""
                        if cashflow.way_to_pay == 'E':
                            payment_type = "EFECTIVO"
                        elif cashflow.way_to_pay == 'Y':
                            payment_type = "YAPE"
                        elif cashflow.way_to_pay == 'D':
                            payment_type = "DEPÓSITO"
                        ws.cell(row=row, column=6, value=payment_type).border = border
                        ws.cell(row=row, column=7, value=decimal.Decimal(cashflow.total)).border = border
                        
                        if i == 0:
                            ws.cell(row=row, column=8, value="PAGADO").border = border
                            ws.cell(row=row, column=9, value=Decimal(data['order'].total)).border = border
                        else:
                            ws.cell(row=row, column=8, value="").border = border
                            ws.cell(row=row, column=9, value="").border = border
                        
                        row += 1
            
            # Totales de ingresos
            row += 1
            ws.cell(row=row, column=8, value="YAPE:").font = Font(bold=True)
            ws.cell(row=row, column=9, value=decimal.Decimal(advances_yape)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=8, value="EFECTIVO:").font = Font(bold=True)
            ws.cell(row=row, column=9, value=decimal.Decimal(advances_efectivo)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=8, value="TOTAL INGRESOS:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=8).fill = header_fill_primary
            ws.cell(row=row, column=9, value=decimal.Decimal(total_advances)).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=9).fill = header_fill_primary
            
            # Sección de SALDOS
            row += 3
            ws.cell(row=row, column=1, value="SALDOS")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_success
            ws.merge_cells(f'A{row}:F{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Encabezados de saldos
            row += 1
            saldos_headers = ['N° COMPROBANTE', 'FECHA', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'S/TOTAL']
            for col, header in enumerate(saldos_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_success
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de saldos
            row += 1
            for cashflow in payments_cashflows:
                ws.cell(row=row, column=1, value=f"{cashflow.order.subsidiary.serial}-{cashflow.order.correlative:03d}").border = border
                ws.cell(row=row, column=2, value=cashflow.order.register_date.strftime('%d-%m-%Y')).border = border
                ws.cell(row=row, column=3, value=cashflow.description or "PAGO TOTAL").border = border
                ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                
                payment_type = ""
                if cashflow.way_to_pay == 'E':
                    payment_type = "EFECTIVO"
                elif cashflow.way_to_pay == 'Y':
                    payment_type = "YAPE"
                elif cashflow.way_to_pay == 'D':
                    payment_type = "DEPÓSITO"
                ws.cell(row=row, column=5, value=payment_type).border = border
                ws.cell(row=row, column=6, value=decimal.Decimal(cashflow.total)).border = border
                row += 1
            
            # Totales de saldos
            row += 1
            ws.cell(row=row, column=5, value="YAPE:").font = Font(bold=True)
            ws.cell(row=row, column=6, value=decimal.Decimal(payments_yape)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=5, value="EFECTIVO:").font = Font(bold=True)
            ws.cell(row=row, column=6, value=decimal.Decimal(payments_efectivo)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=5, value="TOTAL CANCELACIONES:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=5).fill = header_fill_success
            ws.cell(row=row, column=6, value=decimal.Decimal(total_payments)).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=6).fill = header_fill_success
            
            # Sección de EGRESOS
            row += 3
            ws.cell(row=row, column=1, value="EGRESOS")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_danger
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Encabezados de egresos
            row += 1
            egresos_headers = ['NRO', 'DESCRIPCIÓN', 'TIPO EGRESO', 'USUARIO', 'MONTO']
            for col, header in enumerate(egresos_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_danger
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de egresos
            row += 1
            for i, cashflow in enumerate(expenses_cashflows, 1):
                ws.cell(row=row, column=1, value=i).border = border
                ws.cell(row=row, column=2, value=cashflow.description or '-').border = border
                
                expense_type = ""
                if cashflow.type_expense == 'V':
                    expense_type = "VARIABLE"
                elif cashflow.type_expense == 'F':
                    expense_type = "FIJO"
                elif cashflow.type_expense == 'P':
                    expense_type = "PERSONAL"
                elif cashflow.type_expense == 'O':
                    expense_type = "OTRO"
                ws.cell(row=row, column=3, value=expense_type).border = border
                ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                ws.cell(row=row, column=5, value=decimal.Decimal(cashflow.total)).border = border
                row += 1
            
            # Total de egresos
            row += 1
            ws.cell(row=row, column=4, value="TOTAL EGRESOS:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=4).fill = header_fill_danger
            ws.cell(row=row, column=5, value=decimal.Decimal(total_expenses_amount)).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=5).fill = header_fill_danger
            
            # Sección de RESUMENES
            row += 3
            ws.cell(row=row, column=1, value="RESUMENES")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_success
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Resumen de ingresos
            row += 1
            ws.cell(row=row, column=1, value="INGRESOS").font = Font(bold=True, size=12, color="007bff")
            row += 1
            ws.cell(row=row, column=1, value="INGRESOS DEL DÍA:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=decimal.Decimal(total_advances)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value="SALDOS:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=decimal.Decimal(total_payments)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value="SUBTOTAL INGRESOS:").font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_advances + total_payments)).font = Font(bold=True)
            ws.cell(row=row, column=2).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            
            # Resumen de egresos
            row += 2
            ws.cell(row=row, column=1, value="EGRESOS").font = Font(bold=True, size=12, color="dc3545")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL EGRESOS:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=decimal.Decimal(total_expenses_amount)).font = Font(bold=True)
            
            # Resumen final
            row += 3
            ws.cell(row=row, column=1, value="TOTAL EFECTIVO:").font = Font(bold=True, size=11, color="28a745")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_efectivo)).font = Font(bold=True, size=11, color="28a745")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL YAPE:").font = Font(bold=True, size=11, color="17a2b8")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_yape)).font = Font(bold=True, size=11, color="17a2b8")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL EGRESOS:").font = Font(bold=True, size=11, color="dc3545")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_expenses_amount)).font = Font(bold=True, size=11, color="dc3545")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL FINAL:").font = Font(bold=True, size=12, color="ffc107")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_general - total_expenses_amount)).font = Font(bold=True, size=12, color="ffc107")
            ws.cell(row=row, column=2).fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            
            # Ajustar ancho de columnas
            column_widths = [15, 25, 8, 30, 15, 12, 12, 12, 12, 12, 12, 12, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Guardar archivo
            filename = f"reporte_ventas_gastos_{report_date}.xlsx"
            filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            wb.save(filepath)
            
            # Retornar URL del archivo
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte exportado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar el reporte: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)


@csrf_exempt
def export_sales_report_by_user_excel(request):
    """Exportar reporte de ventas por usuario a Excel"""
    if request.method == 'POST':
        try:
            # Obtener datos del reporte
            report_date = request.POST.get('report_date')
            user_id = request.POST.get('user')
            user_obj = None
            
            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por usuario
            if user_id and user_id != '0':
                from apps.users.models import CustomUser
                user_obj = CustomUser.objects.get(id=int(user_id))
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    user_id=user_id
                )
            else:
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']
            ).order_by('id')
            
            if user_id and user_id != '0':
                # Obtener IDs de órdenes que tienen cashflows del usuario del día
                orders_with_user_cashflows = order_cashflows.filter(
                    user_id=user_id
                ).values_list('order_id', flat=True).distinct()
                
                # Filtrar órdenes: creadas por el usuario O con cashflows del usuario
                orders_of_day = orders_of_day.filter(
                    Q(user_id=user_id) | Q(id__in=orders_with_user_cashflows)
                )
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # Crear estructura de datos para adelantos (ingresos del día)
            advances_grouped = {}
            for order in orders_of_day:
                order_cashflows_day = cashflows.filter(
                    order=order,
                    type='E',
                    transaction_date=report_date
                ).order_by('id')
                
                if order_cashflows_day.exists():
                    total_paid = sum(decimal.Decimal(cf.total) for cf in order_cashflows_day)
                    saldo = decimal.Decimal(order.total) - total_paid
                    is_paid_in_full = abs(saldo) < 0.01
                    
                    advances_grouped[order.id] = {
                        'order': order,
                        'cashflows': list(order_cashflows_day),
                        'total_advances': total_paid,
                        'saldo': saldo,
                        'is_paid_in_full': is_paid_in_full,
                        'cashflow_count': order_cashflows_day.count()
                    }
            
            # Preparar datos de saldos (cancelaciones)
            payments_cashflows = order_cashflows.filter(
                type='E',
                order_type_entry='T'
            ).exclude(
                order__register_date__gte=report_date
            )
            
            # Preparar datos de egresos
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'
            )
            
            # Calcular totales
            total_advances = sum(data['total_advances'] for data in advances_grouped.values())
            total_payments = payments_cashflows.aggregate(total=Sum('total'))['total'] or 0
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular totales por tipo de pago
            advances_efectivo = 0
            advances_yape = 0
            advances_deposito = 0
            
            for data in advances_grouped.values():
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        advances_efectivo += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        advances_yape += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'D':
                        advances_deposito += decimal.Decimal(cashflow.total)
            
            payments_efectivo = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
            payments_yape = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
            
            total_efectivo = advances_efectivo + payments_efectivo
            total_yape = advances_yape + payments_yape
            total_general = total_efectivo + total_yape
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Reporte {report_date}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill_primary = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
            header_fill_success = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            header_fill_danger = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            title_font = Font(bold=True, size=16, color="007bff")
            border = Border(
                left=Side(style='medium', color='adb5bd'),
                right=Side(style='medium', color='adb5bd'),
                top=Side(style='medium', color='adb5bd'),
                bottom=Side(style='medium', color='adb5bd')
            )
            
            # Título principal
            ws.merge_cells('A1:J1')
            user_name = f"{user_obj.first_name} {user_obj.last_name}".strip() if user_obj else "TODOS"
            ws['A1'] = f"USUARIO: {user_name.upper()} - DÍA: {datetime.strptime(report_date, '%Y-%m-%d').strftime('%d-%m-%Y')}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Sección de INGRESOS DEL DÍA
            ws['A3'] = "INGRESOS DEL DÍA"
            ws['A3'].font = header_font
            ws['A3'].fill = header_fill_primary
            ws.merge_cells('A3:J3')
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # Encabezados de ingresos
            income_headers = ['N° CPTE.', 'CLIENTE O RAZON SOCIAL', 'CANT.', 'DESCRIPCIÓN DEL PRODUCTO', 'USUARIO', 'TIPO PAGO', 'A CUENTA S/.', 'SALDO S/.', 'TOTAL S/.']
            for col, header in enumerate(income_headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_primary
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de ingresos del día
            row = 5
            for order_id, data in advances_grouped.items():
                if not data['is_paid_in_full']:  # Solo adelantos
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                            ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                            ws.cell(row=row, column=3, value=1).border = border  # Cantidad
                            # Descripción del producto
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            ws.cell(row=row, column=4, value=product_desc).border = border
                        else:
                            # Filas adicionales sin datos de orden
                            ws.cell(row=row, column=1, value="").border = border
                            ws.cell(row=row, column=2, value="").border = border
                            ws.cell(row=row, column=3, value="").border = border
                            ws.cell(row=row, column=4, value="").border = border
                        
                        # Datos del cashflow
                        ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                        
                        # Tipo de pago
                        payment_type = ""
                        if cashflow.way_to_pay == 'E':
                            payment_type = "EFECTIVO"
                        elif cashflow.way_to_pay == 'Y':
                            payment_type = "YAPE"
                        elif cashflow.way_to_pay == 'D':
                            payment_type = "DEPÓSITO"
                        ws.cell(row=row, column=6, value=payment_type).border = border
                        ws.cell(row=row, column=7, value=Decimal(cashflow.total)).border = border
                        
                        if i == 0:  # Solo en la primera fila
                            ws.cell(row=row, column=8, value=Decimal(data['saldo'])).border = border
                            ws.cell(row=row, column=9, value=Decimal(data['order'].total)).border = border
                        else:
                            ws.cell(row=row, column=8, value="").border = border
                            ws.cell(row=row, column=9, value="").border = border
                        
                row += 1
                
                # Pagos completos
                if data['is_paid_in_full']:
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                            ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                            ws.cell(row=row, column=3, value=1).border = border
                            # Descripción del producto
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            ws.cell(row=row, column=4, value=product_desc).border = border
                        else:
                            ws.cell(row=row, column=1, value="").border = border
                            ws.cell(row=row, column=2, value="").border = border
                            ws.cell(row=row, column=3, value="").border = border
                            ws.cell(row=row, column=4, value="").border = border
                        
                        ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                        
                        payment_type = ""
                        if cashflow.way_to_pay == 'E':
                            payment_type = "EFECTIVO"
                        elif cashflow.way_to_pay == 'Y':
                            payment_type = "YAPE"
                        elif cashflow.way_to_pay == 'D':
                            payment_type = "DEPÓSITO"
                        ws.cell(row=row, column=6, value=payment_type).border = border
                        ws.cell(row=row, column=7, value=decimal.Decimal(cashflow.total)).border = border
                        
                        if i == 0:
                            ws.cell(row=row, column=8, value="PAGADO").border = border
                            ws.cell(row=row, column=9, value=Decimal(data['order'].total)).border = border
                        else:
                            ws.cell(row=row, column=8, value="").border = border
                            ws.cell(row=row, column=9, value="").border = border
                        
                        row += 1
            
            # Totales de ingresos
            row += 1
            ws.cell(row=row, column=8, value="YAPE:").font = Font(bold=True)
            ws.cell(row=row, column=9, value=Decimal(advances_yape)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=8, value="EFECTIVO:").font = Font(bold=True)
            ws.cell(row=row, column=9, value=Decimal(advances_efectivo)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=8, value="TOTAL INGRESOS:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=8).fill = header_fill_primary
            ws.cell(row=row, column=9, value=Decimal(total_advances)).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=9).fill = header_fill_primary
            
            # Sección de SALDOS (solo si hay datos)
            if payments_cashflows.exists():
                row += 3  # Espacio entre secciones
                
                # Título de saldos
                ws.cell(row=row, column=1, value="SALDOS")
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=1).fill = header_fill_success
                ws.merge_cells(f'A{row}:F{row}')
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                
                # Encabezados de saldos
                row += 1
                payments_headers = ['N° COMPROBANTE', 'FECHA', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'S/TOTAL']
                for col, header in enumerate(payments_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill_success
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Datos de saldos
                row += 1
                for cashflow in payments_cashflows:
                    ws.cell(row=row, column=1, value=f"{cashflow.order.subsidiary.serial}-{cashflow.order.correlative:03d}").border = border
                    ws.cell(row=row, column=2, value=cashflow.order.register_date.strftime('%d-%m-%Y')).border = border
                    ws.cell(row=row, column=3, value=cashflow.description or "PAGO TOTAL").border = border
                    ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                    
                    # Tipo de pago
                    payment_type = ""
                    if cashflow.way_to_pay == 'E':
                        payment_type = "EFECTIVO"
                    elif cashflow.way_to_pay == 'Y':
                        payment_type = "YAPE"
                    elif cashflow.way_to_pay == 'D':
                        payment_type = "DEPÓSITO"
                    ws.cell(row=row, column=5, value=payment_type).border = border
                    ws.cell(row=row, column=6, value=Decimal(cashflow.total)).border = border
                    row += 1
                
                # Totales de saldos
                payments_efectivo_section = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
                payments_yape_section = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
                
                row += 1
                ws.cell(row=row, column=5, value="YAPE:").font = Font(bold=True)
                ws.cell(row=row, column=6, value=Decimal(payments_yape_section)).font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=5, value="EFECTIVO:").font = Font(bold=True)
                ws.cell(row=row, column=6, value=Decimal(payments_efectivo_section)).font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=5, value="TOTAL CANCELACIONES:").font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=5).fill = header_fill_success
                ws.cell(row=row, column=6, value=Decimal(total_payments)).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=6).fill = header_fill_success
            
            # Sección de EGRESOS (solo si hay datos)
            if expenses_cashflows.exists():
                row += 3  # Espacio entre secciones
                
                # Título de egresos
                ws.cell(row=row, column=1, value="EGRESOS")
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=1).fill = header_fill_danger
                ws.merge_cells(f'A{row}:E{row}')
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                
                # Encabezados de egresos
                row += 1
                expenses_headers = ['NRO', 'DESCRIPCIÓN', 'TIPO EGRESO', 'USUARIO', 'MONTO']
                for col, header in enumerate(expenses_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill_danger
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Datos de egresos
                row += 1
                for i, cashflow in enumerate(expenses_cashflows, 1):
                    ws.cell(row=row, column=1, value=i).border = border
                    ws.cell(row=row, column=2, value=cashflow.description or '-').border = border
                    
                    # Tipo de egreso
                    expense_type = ""
                    if cashflow.type_expense == 'V':
                        expense_type = "VARIABLE"
                    elif cashflow.type_expense == 'F':
                        expense_type = "FIJO"
                    elif cashflow.type_expense == 'P':
                        expense_type = "PERSONAL"
                    elif cashflow.type_expense == 'O':
                        expense_type = "OTRO"
                    ws.cell(row=row, column=3, value=expense_type).border = border
                    ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                    ws.cell(row=row, column=5, value=Decimal(cashflow.total)).border = border
                    row += 1
                
                # Total de egresos
                row += 1
                ws.cell(row=row, column=4, value="TOTAL EGRESOS:").font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=4).fill = header_fill_danger
                ws.cell(row=row, column=5, value=Decimal(total_expenses_amount)).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=5).fill = header_fill_danger
            
            # Sección de RESUMENES
            row += 3  # Espacio entre secciones
            
            # Título de resúmenes
            ws.cell(row=row, column=1, value="RESUMENES")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = PatternFill(start_color="6f42c1", end_color="6f42c1", fill_type="solid")  # Púrpura
            ws.merge_cells(f'A{row}:B{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Calcular totales para resúmenes
            payments_efectivo_total = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
            payments_yape_total = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
            total_general = advances_efectivo + advances_yape + advances_deposito + payments_efectivo_total + payments_yape_total
            
            # Datos de resúmenes
            summary_data = [
                ['CONCEPTO', 'MONTO'],
                ['INGRESOS DEL DÍA:', Decimal(total_advances)],
                ['SALDOS:', Decimal(total_payments)],
                ['SUBTOTAL INGRESOS:', Decimal(total_advances + total_payments)],
                ['TOTAL EGRESOS:', Decimal(total_expenses_amount)],
                ['TOTAL EFECTIVO:', Decimal(advances_efectivo + payments_efectivo_total)],
                ['TOTAL YAPE:', Decimal(advances_yape + payments_yape_total)],
                ['TOTAL FINAL:', Decimal(total_general - total_expenses_amount)]
            ]
            
            row += 1
            for i, (concepto, monto) in enumerate(summary_data):
                if i == 0:  # Encabezado
                    ws.cell(row=row, column=1, value=concepto).font = header_font
                    ws.cell(row=row, column=1).fill = PatternFill(start_color="6f42c1", end_color="6f42c1", fill_type="solid")
                    ws.cell(row=row, column=2, value=monto).font = header_font
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="6f42c1", end_color="6f42c1", fill_type="solid")
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
                else:
                    ws.cell(row=row, column=1, value=concepto).font = Font(bold=True)
                    ws.cell(row=row, column=2, value=monto).font = Font(bold=True)
                    ws.cell(row=row, column=1).border = border
                    ws.cell(row=row, column=2).border = border
                row += 1
            
            # Ajustar ancho de columnas
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Guardar archivo
            filename = f"reporte_ventas_usuario_{report_date}.xlsx"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            wb.save(file_path)
            
            # URL del archivo
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte Excel generado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al generar el Excel: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)
