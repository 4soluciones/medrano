import itertools

from django.shortcuts import render
from django.db.models.functions import Coalesce
from django.shortcuts import render
from django.views.generic import TemplateView, View, CreateView, UpdateView
from django.views.decorators.csrf import csrf_exempt
from django.forms.models import model_to_dict
from django.http import JsonResponse, HttpResponse
from django.views.generic import ListView
from http import HTTPStatus
import re
import locale
import decimal
import calendar

from .models import *
import pytz
from django.contrib.auth.models import User
import json
import requests
import decimal
import math
import random
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models.fields.files import ImageFieldFile
from django.template import loader
from datetime import datetime, date, timedelta
from django.utils import timezone
from django.db import DatabaseError, IntegrityError
from django.core import serializers
from django.db.models import Min, Sum, Max, Q, Prefetch, Subquery, OuterRef, Value, IntegerField, Case, ExpressionWrapper, DecimalField
from django.db.models.functions import ExtractYear
from medrano import settings
import os
from decimal import Decimal
from django.db.models import F


from ..sales.models import Person, Order, OrderDetail
from ..sales.views_API import query_apis_net_money
from ..users.models import CustomUser
from ..hrm.models import Subsidiary


# =============================================================================
# VISTAS PARA GESTIÓN DE CUENTAS/CAJAS
# =============================================================================

def cash_list(request):
    """Vista principal del listado de cuentas/cajas"""
    if request.method == 'GET':
        subsidiary_set = Subsidiary.objects.all()
        currency_types = Cash.CURRENCY_TYPE_CHOICES
        
        return render(request, 'accounting/cash_list.html', {
            'subsidiary_set': subsidiary_set,
            'currency_types': currency_types,
        })
    elif request.method == 'POST':
        try:
            # Filtrar cuentas según parámetros
            subsidiary_id = request.POST.get('subsidiary')
            currency_type = request.POST.get('currency_type')
            
            cash_accounts = Cash.objects.all()
        
            if subsidiary_id and subsidiary_id != '0':
                cash_accounts = cash_accounts.filter(subsidiary_id=subsidiary_id)
            if currency_type and currency_type != '0':
                cash_accounts = cash_accounts.filter(currency_type=currency_type)

            cash_accounts = cash_accounts.select_related('subsidiary').order_by('name')

            tpl = loader.get_template('accounting/cash_list_grid.html')
            context = {
                'cash_accounts': cash_accounts,
            }

            return JsonResponse({
                'grid': tpl.render(context, request),
            }, status=HTTPStatus.OK)
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al cargar las cuentas: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)


def cash_create(request):
    """Vista para crear nueva cuenta/caja"""
    if request.method == 'GET':
        subsidiary_set = Subsidiary.objects.all()
        currency_types = Cash.CURRENCY_TYPE_CHOICES
        
        return render(request, 'accounting/cash_create.html', {
            'subsidiary_set': subsidiary_set,
            'currency_types': currency_types,
        })


@csrf_exempt
def cash_save(request):
    """Vista para guardar nueva cuenta/caja"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            name = request.POST.get('cash_name', '').strip()
            subsidiary_id = request.POST.get('subsidiary_id', '')
            account_number = request.POST.get('account_number', '').strip()
            currency_type = request.POST.get('currency_type', 'S')
            account_type = request.POST.get('account_type', 'C')
            
            # Validaciones básicas
            if not name:
                return JsonResponse({
                    'success': False,
                    'message': 'El nombre de la cuenta es obligatorio'
                }, status=HTTPStatus.BAD_REQUEST)
            
            if not subsidiary_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una sucursal'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Verificar si ya existe una cuenta con el mismo nombre en la misma sucursal
            if Cash.objects.filter(name__iexact=name, subsidiary_id=subsidiary_id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Ya existe una cuenta con el nombre "{name}" en esta sucursal'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Crear la cuenta
            subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
            cash_obj = Cash(
                name=name.upper(),
                subsidiary=subsidiary_obj,
                account_number=account_number.upper() if account_number else None,
                currency_type=currency_type,
                account_type=account_type
            )
            cash_obj.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Cuenta creada exitosamente',
                'cash_id': cash_obj.id
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al crear la cuenta: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


def cash_edit(request, cash_id):
    """Vista para editar cuenta existente"""
    try:
        cash_obj = Cash.objects.select_related('subsidiary').get(id=cash_id)
        subsidiary_set = Subsidiary.objects.all()
        currency_types = Cash.CURRENCY_TYPE_CHOICES
        
        return render(request, 'accounting/cash_edit.html', {
            'cash': cash_obj,
            'subsidiary_set': subsidiary_set,
            'currency_types': currency_types,
        })
        
    except Cash.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Cuenta no encontrada'
        }, status=HTTPStatus.NOT_FOUND)


@csrf_exempt
def cash_get(request):
    """Vista para obtener datos de una cuenta específica"""
    if request.method == 'POST':
        try:
            cash_id = request.POST.get('cash_id')
            if not cash_id:
                return JsonResponse({
                    'success': False,
                    'message': 'ID de cuenta no proporcionado'
                }, status=HTTPStatus.BAD_REQUEST)
            
            cash_obj = Cash.objects.select_related('subsidiary').get(id=int(cash_id))
            
            # Preparar datos para el formulario
            cash_data = {
                'id': cash_obj.id,
                'name': cash_obj.name,
                'subsidiary_id': cash_obj.subsidiary.id,
                'account_number': cash_obj.account_number,
                'currency_type': cash_obj.currency_type,
                'account_type': cash_obj.account_type,
            }
            
            return JsonResponse({
                'success': True,
                'cash': cash_data
            }, status=HTTPStatus.OK)
            
        except Cash.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Cuenta no encontrada'
            }, status=HTTPStatus.NOT_FOUND)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al obtener los datos de la cuenta: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


@csrf_exempt
def cash_update(request):
    """Vista para actualizar cuenta existente"""
    if request.method == 'POST':
        try:
            cash_id = request.POST.get('cash_id')
            if not cash_id:
                return JsonResponse({
                    'success': False,
                    'message': 'ID de cuenta no proporcionado'
                }, status=HTTPStatus.BAD_REQUEST)
            
            cash_obj = Cash.objects.get(id=int(cash_id))
            
            # Obtener datos del formulario
            name = request.POST.get('cash_name', '').strip()
            subsidiary_id = request.POST.get('subsidiary_id', '')
            account_number = request.POST.get('account_number', '').strip()
            currency_type = request.POST.get('currency_type', 'S')
            account_type = request.POST.get('account_type', 'C')
            
            # Validaciones básicas
            if not name:
                return JsonResponse({
                    'success': False,
                    'message': 'El nombre de la cuenta es obligatorio'
                }, status=HTTPStatus.BAD_REQUEST)
            
            if not subsidiary_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una sucursal'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Verificar si ya existe una cuenta con el mismo nombre en la misma sucursal (excluyendo la actual)
            if Cash.objects.filter(name__iexact=name, subsidiary_id=subsidiary_id).exclude(id=cash_obj.id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Ya existe una cuenta con el nombre "{name}" en esta sucursal'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Actualizar la cuenta
            subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
            cash_obj.name = name.upper()
            cash_obj.subsidiary = subsidiary_obj
            cash_obj.account_number = account_number.upper() if account_number else None
            cash_obj.currency_type = currency_type
            cash_obj.account_type = account_type
            cash_obj.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Cuenta actualizada correctamente'
            }, status=HTTPStatus.OK)
            
        except Cash.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Cuenta no encontrada'
            }, status=HTTPStatus.NOT_FOUND)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al actualizar la cuenta: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


@csrf_exempt
def cashflow_get(request):
    """Vista para obtener datos de un gasto específico - Solo para administradores"""
    if request.method == 'POST':
        try:
            # Verificar permisos de administrador
            if not (hasattr(request.user, 'has_access_to_all') and request.user.has_access_to_all):
                return JsonResponse({
                    'success': False,
                    'message': 'No tiene permisos para editar gastos. Solo los administradores pueden realizar esta acción.'
                }, status=HTTPStatus.FORBIDDEN)
            
            cashflow_id = request.POST.get('cashflow_id')
            if not cashflow_id:
                return JsonResponse({
                    'success': False,
                    'message': 'ID de gasto no proporcionado'
                }, status=HTTPStatus.BAD_REQUEST)
            
            cashflow_obj = CashFlow.objects.select_related('cash', 'user').get(id=int(cashflow_id))
            
            # Preparar datos para el formulario
            cashflow_data = {
                'id': cashflow_obj.id,
                'transaction_date': cashflow_obj.transaction_date.strftime('%Y-%m-%d'),
                'type': cashflow_obj.type,
                'cash_id': cashflow_obj.cash.id,
                'type_expense': cashflow_obj.type_expense,
                'user_id': cashflow_obj.user.id,
                'document_type_attached': cashflow_obj.document_type_attached,
                'description': cashflow_obj.description,
                'serial': cashflow_obj.serial,
                'n_receipt': cashflow_obj.n_receipt,
                'operation_code': cashflow_obj.operation_code,
                'subtotal': float(cashflow_obj.subtotal),
                'igv': float(cashflow_obj.igv),
                'total': float(cashflow_obj.total),
            }
            
            return JsonResponse({
                'success': True,
                'cashflow': cashflow_data
            }, status=HTTPStatus.OK)
            
        except CashFlow.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Gasto no encontrado'
            }, status=HTTPStatus.NOT_FOUND)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al obtener los datos del gasto: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


@csrf_exempt
def cashflow_delete(request):
    """Vista para eliminar gasto existente - Solo para administradores"""
    if request.method == 'POST':
        try:
            # Verificar permisos de administrador
            if not (hasattr(request.user, 'has_access_to_all') and request.user.has_access_to_all):
                return JsonResponse({
                    'success': False,
                    'message': 'No tiene permisos para eliminar gastos. Solo los administradores pueden realizar esta acción.'
                }, status=HTTPStatus.FORBIDDEN)
            
            cashflow_id = request.POST.get('cashflow_id')
            if not cashflow_id:
                return JsonResponse({
                    'success': False,
                    'message': 'ID de gasto no proporcionado'
                }, status=HTTPStatus.BAD_REQUEST)
            
            cashflow_obj = CashFlow.objects.get(id=int(cashflow_id))
            
            # Obtener información del gasto para el mensaje
            description = cashflow_obj.description
            amount = cashflow_obj.total
            
            # Eliminar el gasto
            cashflow_obj.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Gasto eliminado exitosamente: {description} - S/ {amount}'
            }, status=HTTPStatus.OK)
            
        except CashFlow.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Gasto no encontrado'
            }, status=HTTPStatus.NOT_FOUND)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al eliminar el gasto: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


# =============================================================================
# VISTAS PARA GESTIÓN DE GASTOS (CASHFLOW)
# =============================================================================

def cashflow_list(request):
    """Vista principal del listado de gastos"""
    if request.method == 'GET':
        document_types = CashFlow.DOCUMENT_TYPE_ATTACHED_CHOICES
        transaction_types = [('A', 'Apertura'), ('C', 'Cierre'), ('E', 'Entrada'), ('S', 'Salida')]  # Apertura, cierre, entrada y salida
        expense_types = CashFlow.TYPE_EXPENSE
        user_set = CustomUser.objects.filter(is_active=True, is_staff=False)
        subsidiary_set = Subsidiary.objects.all()
        # Fecha actual para los filtros
        peru_tz = pytz.timezone("America/Lima")
        date_now = datetime.now(peru_tz).strftime('%Y-%m-%d')
        
        # Obtener la sucursal del usuario actual
        user_subsidiary = None
        first_cash_account = None

        if hasattr(request.user, 'subsidiary') and request.user.subsidiary:
            user_subsidiary = request.user.subsidiary

        # Verificar si el usuario tiene permisos de administrador
        is_admin = hasattr(request.user, 'has_access_to_all') and request.user.has_access_to_all

        # Filtrar cajas según permisos del usuario
        if is_admin:
            # Usuario admin puede ver todas las cajas
            cash_accounts = Cash.objects.all()
            # Buscar la primera cuenta de tipo 'E' de cualquier sucursal
            first_cash_account = Cash.objects.filter(account_type='E').first()
        else:
            # Usuario normal solo ve cajas de su sucursal
            cash_accounts = Cash.objects.filter(subsidiary=user_subsidiary)
            # Buscar la primera cuenta de tipo 'E' de la sucursal del usuario
            if user_subsidiary:
                first_cash_account = Cash.objects.filter(
                    subsidiary=user_subsidiary,
                    account_type='E'
                ).first()

        # Si no hay cuenta de tipo 'E', buscar cualquier cuenta según permisos
        if not first_cash_account:
            if is_admin:
                first_cash_account = Cash.objects.first()
            elif user_subsidiary:
                first_cash_account = Cash.objects.filter(subsidiary=user_subsidiary).first()
        
        return render(request, 'accounting/cashflow_list.html', {
            'cash_accounts': cash_accounts,
            'document_types': document_types,
            'transaction_types': transaction_types,
            'expense_types': expense_types,
            'subsidiary_set': subsidiary_set,
            'user_set': user_set,
            'date_now': date_now,
            'user_subsidiary': user_subsidiary,
            'first_cash_account': first_cash_account,
            'is_admin': is_admin,
        })
    elif request.method == 'POST':
        try:
            # Filtrar gastos según parámetros
            cash_id = request.POST.get('cash_account')
            report_date = request.POST.get('report_date')
            
            cashflows = CashFlow.objects.filter(type__in=['S', 'A'])
        
            if cash_id and cash_id != '0':
                cashflows = cashflows.filter(cash_id=cash_id)

            # Filtro de fecha única
            if report_date:
                cashflows = cashflows.filter(transaction_date=report_date)

            # Ordenar: aperturas primero, luego por id
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary').order_by(
                'type',  # Aperturas (A) aparecerán primero
                'id'
            )

            # Calcular totales
            # Entradas: tipo 'E' (Entrada) + tipo 'A' (Apertura)
            total_income = cashflows.filter(type__in=['E', 'A']).aggregate(total=Sum('total'))['total'] or 0
            # Salidas: tipo 'S' (Salida)
            total_expenses = cashflows.filter(type='S').aggregate(total=Sum('total'))['total'] or 0
            # Balance: Entradas - Salidas
            net_balance = total_income - total_expenses

            # Totales por tipo de gasto
            expense_totals = {}
            for expense_code, expense_name in CashFlow.TYPE_EXPENSE:
                total = cashflows.filter(type__in=['S', 'A'], type_expense=expense_code).aggregate(
                    total=Sum('total')
                )['total'] or 0
                expense_totals[expense_code] = total

            tpl = loader.get_template('accounting/cashflow_list_grid.html')
            context = {
                'cashflows': cashflows,
                'total_income': total_income,
                'total_expenses': total_expenses,
                'net_balance': net_balance,
                'expense_totals': expense_totals,
                # 'date_now': date_now,
            }

            return JsonResponse({
                'grid': tpl.render(context, request),
            }, status=HTTPStatus.OK)
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al cargar los gastos: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)


def cashflow_create(request):
    """Vista para crear nuevo gasto"""
    if request.method == 'GET':
        cash_accounts = Cash.objects.all()
        document_types = CashFlow.DOCUMENT_TYPE_ATTACHED_CHOICES
        transaction_types = [('A', 'Apertura'), ('C', 'Cierre'), ('E', 'Entrada'), ('S', 'Salida')]  # Apertura, cierre, entrada y salida
        expense_types = CashFlow.TYPE_EXPENSE
        user_set = CustomUser.objects.filter(is_active=True, is_staff=False)
        
        # Fecha actual
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        return render(request, 'accounting/cashflow_create.html', {
            'cash_accounts': cash_accounts,
            'document_types': document_types,
            'transaction_types': transaction_types,
            'expense_types': expense_types,
            'user_set': user_set,
            'date_now': date_now,
        })


@csrf_exempt
def cashflow_save(request):
    """Vista para guardar nuevo gasto"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            transaction_date = request.POST.get('transaction_date')
            description = request.POST.get('description', '').strip()
            serial = request.POST.get('serial', '').strip()
            n_receipt = request.POST.get('n_receipt', '0')
            document_type = request.POST.get('document_type', 'O')
            transaction_type = request.POST.get('transaction_type', 'S')
            subtotal = request.POST.get('subtotal', '0.00')
            total = request.POST.get('total', '0.00')
            igv = request.POST.get('igv', '0.00')
            cash_id = request.POST.get('cash_id')
            operation_code = request.POST.get('operation_code', '').strip()
            expense_type = request.POST.get('expense_type', 'O')
            user_id = request.POST.get('user_id')
            
            # Validaciones básicas
            if not transaction_date:
                return JsonResponse({
                    'success': False,
                    'message': 'La fecha de transacción es obligatoria'
                }, status=HTTPStatus.BAD_REQUEST)
            
            if not description:
                return JsonResponse({
                    'success': False,
                    'message': 'La descripción es obligatoria'
                }, status=HTTPStatus.BAD_REQUEST)
            
            if not cash_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una cuenta/caja'
                }, status=HTTPStatus.BAD_REQUEST)
            
            if not user_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar un usuario'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Obtener objetos relacionados
            cash_obj = Cash.objects.get(id=int(cash_id))
            user_obj = CustomUser.objects.get(id=int(user_id))
            
            # Crear el gasto
            cashflow_obj = CashFlow(
                transaction_date=transaction_date,
                description=description.upper(),
                serial=serial.upper() if serial else None,
                n_receipt=int(n_receipt) if n_receipt else 0,
                document_type_attached=document_type,
                type=transaction_type,
                subtotal=Decimal(str(subtotal)),
                total=Decimal(str(total)),
                igv=Decimal(str(igv)),
                cash=cash_obj,
                operation_code=operation_code.upper() if operation_code else None,
                type_expense=expense_type,
                user=user_obj,
                subsidiary=cash_obj.subsidiary
            )
            cashflow_obj.save()
            
            return JsonResponse({
                'success': True,
                'message': f'{cashflow_obj.get_type_display()} registrado exitosamente',
                'cashflow_id': cashflow_obj.id
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al registrar el gasto: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


def cashflow_edit(request, cashflow_id):
    """Vista para editar gasto existente"""
    try:
        cashflow_obj = CashFlow.objects.select_related('cash', 'user', 'cash__subsidiary').get(id=cashflow_id)
        cash_accounts = Cash.objects.all()
        document_types = CashFlow.DOCUMENT_TYPE_ATTACHED_CHOICES
        transaction_types = [('A', 'Apertura'), ('C', 'Cierre'), ('E', 'Entrada'), ('S', 'Salida')]  # Apertura, cierre, entrada y salida
        expense_types = CashFlow.TYPE_EXPENSE
        user_set = CustomUser.objects.filter(is_active=True, is_staff=False)
        
        return render(request, 'accounting/cashflow_edit.html', {
            'cashflow': cashflow_obj,
            'cash_accounts': cash_accounts,
            'document_types': document_types,
            'transaction_types': transaction_types,
            'expense_types': expense_types,
            'user_set': user_set,
        })
        
    except CashFlow.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Gasto no encontrado'
        }, status=HTTPStatus.NOT_FOUND)


@csrf_exempt
def cashflow_update(request):
    """Vista para actualizar gasto existente"""
    if request.method == 'POST':
        try:
            cashflow_id = request.POST.get('cashflow_id')
            if not cashflow_id:
                return JsonResponse({
                    'success': False,
                    'message': 'ID de gasto no proporcionado'
                }, status=HTTPStatus.BAD_REQUEST)
            
            cashflow_obj = CashFlow.objects.get(id=int(cashflow_id))
            
            # Obtener datos del formulario
            transaction_date = request.POST.get('transaction_date')
            description = request.POST.get('description', '').strip()
            serial = request.POST.get('serial', '').strip()
            n_receipt = request.POST.get('n_receipt', '0')
            document_type = request.POST.get('document_type', 'O')
            transaction_type = request.POST.get('transaction_type', 'S')
            subtotal = request.POST.get('subtotal', '0.00')
            total = request.POST.get('total', '0.00')
            igv = request.POST.get('igv', '0.00')
            cash_id = request.POST.get('cash_id')
            operation_code = request.POST.get('operation_code', '').strip()
            expense_type = request.POST.get('expense_type', 'O')
            user_id = request.POST.get('user_id')
            
            # Validaciones básicas
            if not transaction_date:
                return JsonResponse({
                    'success': False,
                    'message': 'La fecha de transacción es obligatoria'
                }, status=HTTPStatus.BAD_REQUEST)
            
            if not description:
                return JsonResponse({
                    'success': False,
                    'message': 'La descripción es obligatoria'
                }, status=HTTPStatus.BAD_REQUEST)
            
            if not cash_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una cuenta/caja'
                }, status=HTTPStatus.BAD_REQUEST)
            
            if not user_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar un usuario'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Obtener objetos relacionados
            cash_obj = Cash.objects.get(id=int(cash_id))
            user_obj = CustomUser.objects.get(id=int(user_id))
            
            # Actualizar el gasto
            cashflow_obj.transaction_date = transaction_date
            cashflow_obj.description = description.upper()
            cashflow_obj.serial = serial.upper() if serial else None
            cashflow_obj.n_receipt = int(n_receipt) if n_receipt else 0
            cashflow_obj.document_type_attached = document_type
            cashflow_obj.type = transaction_type
            cashflow_obj.subtotal = Decimal(str(subtotal))
            cashflow_obj.total = Decimal(str(total))
            cashflow_obj.igv = Decimal(str(igv))
            cashflow_obj.cash = cash_obj
            cashflow_obj.operation_code = operation_code.upper() if operation_code else None
            cashflow_obj.type_expense = expense_type
            cashflow_obj.user = user_obj
            cashflow_obj.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Gasto actualizado correctamente'
            }, status=HTTPStatus.OK)
            
        except CashFlow.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Gasto no encontrado'
            }, status=HTTPStatus.NOT_FOUND)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al actualizar el gasto: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


def get_cash_accounts_by_subsidiary(request):
    """Vista para obtener cuentas por sucursal"""
    if request.method == 'GET':
        subsidiary_id = request.GET.get('subsidiary', '')
        if subsidiary_id:
            try:
                cash_accounts = Cash.objects.filter(subsidiary_id=int(subsidiary_id)).order_by('name')
                accounts_list = []
                
                for account in cash_accounts:
                    accounts_list.append({
                        'id': account.id,
                        'name': account.name,
                        'currency': account.get_currency_type_display(),
                        'account_type': account.account_type,
                        # 'balance': float(account.initial)
                    })
                
                return JsonResponse({
                    'success': True,
                    'accounts': accounts_list
                }, status=HTTPStatus.OK)
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error al obtener las cuentas: {str(e)}'
                }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
        
        return JsonResponse({
            'success': False,
            'message': 'ID de sucursal no proporcionado'
        }, status=HTTPStatus.BAD_REQUEST)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


# =============================================================================
# VISTAS PARA REPORTES
# =============================================================================
@csrf_exempt
def monthly_report(request):
    """Vista para reporte mensual con gráficos profesionales"""
    if request.method == 'GET':
        subsidiary_set = Subsidiary.objects.all()
        
        # Obtener la sucursal del usuario actual
        user_subsidiary = None
        if hasattr(request.user, 'subsidiary') and request.user.subsidiary:
            user_subsidiary = request.user.subsidiary
        
        # Fecha actual para el filtro
        current_date = datetime.now()
        current_month = current_date.strftime('%Y-%m')
        
        return render(request, 'accounting/monthly_report.html', {
            'subsidiary_set': subsidiary_set,
            'user_subsidiary': user_subsidiary,
            'current_month': current_month,
        })
    elif request.method == 'POST':
        try:
            # Obtener parámetros del filtro
            report_month = request.POST.get('report_month')
            subsidiary_id = request.POST.get('subsidiary')
            
            if not report_month:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar un mes'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Convertir mes a rango de fechas
            year, month = report_month.split('-')
            start_date = datetime(int(year), int(month), 1)
            if int(month) == 12:
                end_date = datetime(int(year) + 1, 1, 1)
            else:
                end_date = datetime(int(year), int(month) + 1, 1)
            
            # Filtrar datos por sucursal si se especifica
            subsidiary_obj = None
            if subsidiary_id and subsidiary_id != '0':
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                orders_filter = {'subsidiary_id': subsidiary_id}
                cashflows_filter = {'cash__subsidiary_id': subsidiary_id}
            else:
                orders_filter = {}
                cashflows_filter = {}
            
            # 1. Órdenes completadas del mes
            completed_orders = Order.objects.filter(
                register_date__gte=start_date,
                register_date__lt=end_date,
                status='C',
                **orders_filter
            ).select_related('subsidiary', 'client')
            
            # 2. Órdenes pendientes por sucursal y en general
            pending_orders = Order.objects.filter(
                register_date__gte=start_date,
                register_date__lt=end_date,
                status='P',
                **orders_filter
            ).select_related('subsidiary', 'client')
            
            # 3. Productos más solicitados
            from django.db.models import Sum, Count
            top_products = OrderDetail.objects.filter(
                order__register_date__gte=start_date,
                order__register_date__lt=end_date,
                order__status__in=['P', 'C'],
                **{'order__' + k: v for k, v in orders_filter.items()}
            ).values('product__name').annotate(
                total_quantity=Sum('quantity'),
                total_orders=Count('order')
            ).order_by('-total_quantity')[:10]
            
            # 4. Órdenes pendientes de entrega
            pending_delivery = Order.objects.filter(
                register_date__gte=start_date,
                register_date__lt=end_date,
                delivery_status='P',
                status__in=['P', 'C'],
                **orders_filter
            ).select_related('subsidiary', 'client')
            
            # 5. Ingresos mensuales (CashFlow tipo 'E')
            monthly_income = CashFlow.objects.filter(
                transaction_date__gte=start_date,
                transaction_date__lt=end_date,
                type='E',
                **cashflows_filter
            ).select_related('cash', 'cash__subsidiary')
            
            # 6. Gastos por sucursal
            monthly_expenses = CashFlow.objects.filter(
                transaction_date__gte=start_date,
                transaction_date__lt=end_date,
                type='S',
                **cashflows_filter
            ).select_related('cash', 'cash__subsidiary')
            
            # Calcular estadísticas
            stats = {
                'completed_orders_count': completed_orders.count(),
                'completed_orders_total': completed_orders.aggregate(Sum('total'))['total__sum'] or 0,
                'pending_orders_count': pending_orders.count(),
                'pending_orders_total': pending_orders.aggregate(Sum('total'))['total__sum'] or 0,
                'pending_delivery_count': pending_delivery.count(),
                'monthly_income_total': monthly_income.aggregate(Sum('total'))['total__sum'] or 0,
                'monthly_expenses_total': monthly_expenses.aggregate(Sum('total'))['total__sum'] or 0,
            }
            
            # Datos para gráficos
            chart_data = {
                'completed_orders_by_subsidiary': list(
                    completed_orders.values('subsidiary__name')
                    .annotate(count=Count('id'), total=Sum('total'))
                    .order_by('-count')
                ),
                'pending_orders_by_subsidiary': list(
                    pending_orders.values('subsidiary__name')
                    .annotate(count=Count('id'), total=Sum('total'))
                    .order_by('-count')
                ),
                'top_products': list(top_products),
                'income_by_subsidiary': list(
                    monthly_income.values('cash__subsidiary__name')
                    .annotate(total=Sum('total'))
                    .order_by('-total')
                ),
                'expenses_by_subsidiary': list(
                    monthly_expenses.values('cash__subsidiary__name')
                    .annotate(total=Sum('total'))
                    .order_by('-total')
                ),
                'daily_completed_orders': list(
                    completed_orders.extra(
                        select={'day': 'DATE(register_date)'}
                    ).values('day').annotate(count=Count('id')).order_by('day')
                ),
                'daily_income': list(
                    monthly_income.extra(
                        select={'day': 'DATE(transaction_date)'}
                    ).values('day').annotate(total=Sum('total')).order_by('day')
                ),
            }
            
            return JsonResponse({
                'success': True,
                'stats': stats,
                'chart_data': chart_data,
                'subsidiary': subsidiary_obj.name if subsidiary_obj else 'Todas las sucursales',
                'month_name': start_date.strftime('%B %Y')
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al generar reporte mensual: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)

@csrf_exempt
def weekly_report(request):
    from datetime import datetime
    """Vista para reporte semanal con gráficos profesionales"""
    if request.method == 'GET':
        subsidiary_set = Subsidiary.objects.all()
        
        # Obtener la sucursal del usuario actual
        user_subsidiary = None
        if hasattr(request.user, 'subsidiary') and request.user.subsidiary:
            user_subsidiary = request.user.subsidiary
        
        # Fecha actual para el filtro
        current_date = datetime.now()
        current_week = current_date.isocalendar()
        current_week_str = f"{current_week[0]}-W{current_week[1]:02d}"
        
        return render(request, 'accounting/weekly_report.html', {
            'subsidiary_set': subsidiary_set,
            'user_subsidiary': user_subsidiary,
            'current_week': current_week_str,
        })
    elif request.method == 'POST':
        try:
            # Obtener parámetros del filtro
            report_week = request.POST.get('report_week')
            subsidiary_id = request.POST.get('subsidiary')
            
            if not report_week:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una semana'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Convertir semana a rango de fechas
            year, week = report_week.split('-W')
            year = int(year)
            week = int(week)
            
            # Calcular fechas de inicio y fin de la semana
            from datetime import datetime, timedelta
            jan_1 = datetime(year, 1, 1)
            start_date = jan_1 + timedelta(weeks=week-1, days=-jan_1.weekday())
            end_date = start_date + timedelta(days=7)
            
            # Filtrar datos por sucursal si se especifica
            subsidiary_obj = None
            if subsidiary_id and subsidiary_id != '0':
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                orders_filter = {'subsidiary_id': subsidiary_id}
                cashflows_filter = {'cash__subsidiary_id': subsidiary_id}
            else:
                orders_filter = {}
                cashflows_filter = {}
            
            # 1. Órdenes completadas de la semana
            completed_orders = Order.objects.filter(
                register_date__gte=start_date,
                register_date__lt=end_date,
                status='C',
                **orders_filter
            ).select_related('subsidiary', 'client')
            
            # 2. Órdenes pendientes por sucursal y en general
            pending_orders = Order.objects.filter(
                register_date__gte=start_date,
                register_date__lt=end_date,
                status='P',
                **orders_filter
            ).select_related('subsidiary', 'client')
            
            # 3. Productos más solicitados
            from django.db.models import Sum, Count
            top_products = OrderDetail.objects.filter(
                order__register_date__gte=start_date,
                order__register_date__lt=end_date,
                order__status__in=['P', 'C'],
                **{'order__' + k: v for k, v in orders_filter.items()}
            ).values('product__name').annotate(
                total_quantity=Sum('quantity'),
                total_orders=Count('order')
            ).order_by('-total_quantity')[:10]
            
            # 4. Órdenes pendientes de entrega
            pending_delivery = Order.objects.filter(
                register_date__gte=start_date,
                register_date__lt=end_date,
                delivery_status='P',
                status__in=['P', 'C'],
                **orders_filter
            ).select_related('subsidiary', 'client')
            
            # 5. Ingresos semanales (CashFlow tipo 'E')
            weekly_income = CashFlow.objects.filter(
                transaction_date__gte=start_date,
                transaction_date__lt=end_date,
                type='E',
                **cashflows_filter
            ).select_related('cash', 'cash__subsidiary')
            
            # 6. Gastos por sucursal
            weekly_expenses = CashFlow.objects.filter(
                transaction_date__gte=start_date,
                transaction_date__lt=end_date,
                type='S',
                **cashflows_filter
            ).select_related('cash', 'cash__subsidiary')
            
            # Calcular estadísticas
            stats = {
                'completed_orders_count': completed_orders.count(),
                'completed_orders_total': completed_orders.aggregate(Sum('total'))['total__sum'] or 0,
                'pending_orders_count': pending_orders.count(),
                'pending_orders_total': pending_orders.aggregate(Sum('total'))['total__sum'] or 0,
                'pending_delivery_count': pending_delivery.count(),
                'weekly_income_total': weekly_income.aggregate(Sum('total'))['total__sum'] or 0,
                'weekly_expenses_total': weekly_expenses.aggregate(Sum('total'))['total__sum'] or 0,
            }
            
            # Datos para gráficos
            chart_data = {
                'completed_orders_by_subsidiary': list(
                    completed_orders.values('subsidiary__name')
                    .annotate(count=Count('id'), total=Sum('total'))
                    .order_by('-count')
                ),
                'pending_orders_by_subsidiary': list(
                    pending_orders.values('subsidiary__name')
                    .annotate(count=Count('id'), total=Sum('total'))
                    .order_by('-count')
                ),
                'top_products': list(top_products),
                'income_by_subsidiary': list(
                    weekly_income.values('cash__subsidiary__name')
                    .annotate(total=Sum('total'))
                    .order_by('-total')
                ),
                'expenses_by_subsidiary': list(
                    weekly_expenses.values('cash__subsidiary__name')
                    .annotate(total=Sum('total'))
                    .order_by('-total')
                ),
                'daily_completed_orders': list(
                    completed_orders.extra(
                        select={'day': 'DATE(register_date)'}
                    ).values('day').annotate(count=Count('id')).order_by('day')
                ),
                'daily_income': list(
                    weekly_income.extra(
                        select={'day': 'DATE(transaction_date)'}
                    ).values('day').annotate(total=Sum('total')).order_by('day')
                ),
            }
            
            return JsonResponse({
                'success': True,
                'stats': stats,
                'chart_data': chart_data,
                'subsidiary': subsidiary_obj.name if subsidiary_obj else 'Todas las sucursales',
                'week_name': f'Semana {week} de {year}'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al generar reporte semanal: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)


def sales_report_test(request):
    """Vista de prueba para el reporte"""
    if request.method == 'GET':
        subsidiary_set = Subsidiary.objects.all()
        cash_accounts = Cash.objects.all()
        
        # Fecha actual para el filtro
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        return render(request, 'accounting/sales_report_test.html', {
            'subsidiary_set': subsidiary_set,
            'cash_accounts': cash_accounts,
            'date_now': date_now,
        })


def sales_report(request):
    """Vista principal del reporte de ventas y gastos"""
    if request.method == 'GET':
        subsidiary_set = Subsidiary.objects.all()
        cash_accounts = Cash.objects.all()
        
        # Fecha actual para el filtro
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        # Obtener la sucursal del usuario actual
        user_subsidiary = None
        first_cash_account = None
        
        if hasattr(request.user, 'subsidiary') and request.user.subsidiary:
            user_subsidiary = request.user.subsidiary
            # Buscar la primera cuenta de tipo 'C' (Caja Chica/Efectivo) de la sucursal del usuario
            first_cash_account = Cash.objects.filter(
                subsidiary=user_subsidiary,
                account_type='C'
            ).first()
        
        # Si no hay cuenta de tipo 'C', buscar cualquier cuenta de la sucursal
        if not first_cash_account and user_subsidiary:
            first_cash_account = Cash.objects.filter(subsidiary=user_subsidiary).first()
        
        return render(request, 'accounting/sales_report.html', {
            'subsidiary_set': subsidiary_set,
            'cash_accounts': cash_accounts,
            'date_now': date_now,
            'user_subsidiary': user_subsidiary,
            'first_cash_account': first_cash_account,
        })
    elif request.method == 'POST':
        try:
            # Obtener parámetros del filtro
            report_date = request.POST.get('report_date')
            subsidiary_id = request.POST.get('subsidiary')
            cash_id = request.POST.get('cash_account')
            subsidiary_obj = None

            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por sucursal
            if subsidiary_id and subsidiary_id != '0':
                # Filtrar por sucursal del usuario logueado
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    cash__subsidiary_id=subsidiary_id
                )
            else:
                # Si no hay sucursal específica, mostrar todos los cashflows del día
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            # Excluir órdenes anuladas (status='A')
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Calcular totales desde accounting_cashflow
            # total_sales: Suma total de todas las órdenes relacionadas con cashflows del día
            total_sales = sum(cf.order.total for cf in order_cashflows if cf.order) or 0
            
            # total_cash_advance: Suma de adelantos (tipo 'A') desde accounting_cashflow
            total_cash_advance = order_cashflows.filter(
                type='E',  # Entrada
                order_type_entry='A'  # Adelanto
            ).aggregate(total=Sum('total'))['total'] or 0
            
            # total_paid: Suma de pagos totales (tipo 'T') desde accounting_cashflow
            total_paid = order_cashflows.filter(
                type='E',  # Entrada
                order_type_entry='T'  # Pago Total
            ).aggregate(total=Sum('total'))['total'] or 0
            
            # total_balance: Saldo pendiente (ventas totales - adelantos totales - pagos totales)
            total_balance = total_sales - total_cash_advance - total_paid
            
            # Calcular totales de gastos
            total_income = cashflows.filter(type='E').aggregate(total=Sum('total'))['total'] or 0
            total_expenses = cashflows.filter(type='S').aggregate(total=Sum('total'))['total'] or 0
            net_expenses = total_expenses - total_income
            
            # Calcular subtotales por tipo de gasto según TYPE_EXPENSE
            total_variable_expenses = cashflows.filter(type='S', type_expense='V').aggregate(total=Sum('total'))['total'] or 0
            total_fixed_expenses = cashflows.filter(type='S', type_expense='F').aggregate(total=Sum('total'))['total'] or 0
            total_personal_expenses = cashflows.filter(type='S', type_expense='P').aggregate(total=Sum('total'))['total'] or 0
            total_other_expenses = cashflows.filter(type='S', type_expense='O').aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular ingreso real: A Cuenta + Total Pagado
            real_income = total_cash_advance + total_paid
            
            # Calcular caja final
            # final_cash: Caja final = Ingreso real - Gastos netos
            # Donde: Gastos netos = Gastos totales - Ingresos por otros conceptos
            final_cash = real_income - net_expenses

            # NUEVA LÓGICA: 
            # 1. ADELANTOS = Todas las órdenes del día (con sus cashflows)
            # 2. SALDOS = Solo cashflows de cancelación (order_type_entry='T')
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']  # Pendientes y completadas
            ).order_by('id')
            
            if subsidiary_id and subsidiary_id != '0':
                orders_of_day = orders_of_day.filter(subsidiary_id=subsidiary_id)
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # Crear estructura de datos más clara para el template
            advances_grouped = {}
            for order in orders_of_day:
                # Obtener solo los cashflows de ADELANTOS relacionados con esta orden del día del reporte
                order_cashflows_day = cashflows.filter(
                    order=order,
                    type='E',  # Solo entradas
                    order_type_entry='A',  # Solo adelantos (no cancelaciones)
                    transaction_date=report_date  # Solo del día del reporte
                ).order_by('id')
                
                if order_cashflows_day.exists():
                    total_paid = sum(float(cf.total) for cf in order_cashflows_day)
                    saldo = float(order.total) - total_paid
                    
                    # Determinar si la orden se pagó en su totalidad (con tolerancia de 1 céntimo)
                    is_paid_in_full = abs(saldo) < 0.01
                    
                    # Crear estructura con información de la orden y sus cashflows
                    advances_grouped[order.id] = {
                        'order': order,
                        'cashflows': list(order_cashflows_day),  # Lista de cashflows individuales
                        'total_advances': total_paid,
                        'saldo': saldo,
                        'is_paid_in_full': is_paid_in_full,
                        'cashflow_count': order_cashflows_day.count()  # Cantidad de cashflows
                    }
            
            # Preparar datos de saldos (solo cashflows de cancelación - order_type_entry='T')
            # EXCLUIR órdenes que fueron creadas el mismo día del reporte (esas van en ingresos del día)
            payments_cashflows = order_cashflows.filter(
                type='E',  # Solo entradas
                order_type_entry='T'  # Solo pagos totales (cancelaciones)
            ).exclude(
                order__register_date__gte=report_date  # Excluir órdenes del mismo día o posteriores
            )
            
            # Preparar datos de cashflows sin order_id (egresos)
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'  # Solo salidas (gastos)
            )

            # Calcular totales para resúmenes
            # Total de ingresos del día (suma de todos los cashflows de las órdenes del día)
            total_advances = sum(data['total_advances'] for data in advances_grouped.values())
            total_payments = payments_cashflows.aggregate(total=Sum('total'))['total'] or 0
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular total de apertura de caja (tipo 'A')
            total_apertura = cashflows.filter(type='A').aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular totales por tipo de pago para adelantos (ingresos del día)
            advances_efectivo = 0
            advances_yape = 0
            advances_deposito = 0
            
            for data in advances_grouped.values():
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        advances_efectivo += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        advances_yape += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'D':
                        advances_deposito += decimal.Decimal(cashflow.total)
            
            payments_efectivo = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
            payments_yape = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
            payments_deposito = payments_cashflows.filter(way_to_pay='D').aggregate(total=Sum('total'))['total'] or 0
            
            # Totales generales
            total_efectivo = advances_efectivo + payments_efectivo
            total_yape = advances_yape + payments_yape
            total_deposito = advances_deposito + payments_deposito
            total_general = total_efectivo + total_yape + total_deposito + total_apertura

            context = {
                'report_date': datetime.strptime(report_date, "%Y-%m-%d").strftime("%d-%m-%Y"),
                'advances_grouped': advances_grouped,
                'orders_of_day': orders_of_day,  # Todas las órdenes del día
                'payments_cashflows': payments_cashflows.order_by('order_id'),  # Solo cancelaciones
                'expenses_cashflows': expenses_cashflows.order_by('id'),
                'total_advances': total_advances,  # Total ingresos del día
                'total_payments': total_payments,  # Total cancelaciones
                'total_expenses_amount': total_expenses_amount,
                'total_apertura': total_apertura,  # Total apertura de caja
                'advances_efectivo': advances_efectivo,
                'advances_yape': advances_yape,
                'advances_deposito': advances_deposito,
                'payments_efectivo': payments_efectivo,
                'payments_yape': payments_yape,
                'payments_deposito': payments_deposito,
                'total_efectivo': total_efectivo,
                'total_yape': total_yape,
                'total_deposito': total_deposito,
                'total_general': total_general,
                'subsidiary': subsidiary_obj,
            }
            
            tpl = loader.get_template('accounting/sales_report_grid.html')
            
            return JsonResponse({
                'success': True,
                'grid': tpl.render(context, request),
                'summary': {
                    'total_sales': float(total_sales),
                    'total_cash_advance': float(total_cash_advance),
                    'total_paid': float(total_paid),
                    'total_balance': float(total_balance),
                    'real_income': float(real_income),
                    'total_income': float(total_income),
                    'total_expenses': float(total_expenses),
                    'net_expenses': float(net_expenses),
                    'final_cash': float(final_cash),
                    'total_variable_expenses': float(total_variable_expenses),
                    'total_fixed_expenses': float(total_fixed_expenses),
                    'total_personal_expenses': float(total_personal_expenses),
                    'total_other_expenses': float(total_other_expenses),
                }
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            print(f"ERROR en sales_report: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'message': f'Error al generar el reporte: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)


@csrf_exempt
def export_sales_report_excel(request):
    """Exportar reporte de ventas a Excel"""
    if request.method == 'POST':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Obtener datos del reporte
            report_date = request.POST.get('report_date')
            subsidiary_id = request.POST.get('subsidiary')
            cash_id = request.POST.get('cash_account')
            subsidiary_obj = None
            
            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por sucursal
            if subsidiary_id and subsidiary_id != '0':
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    cash__subsidiary_id=subsidiary_id
                )
            else:
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']
            ).order_by('id')
            
            if subsidiary_id and subsidiary_id != '0':
                orders_of_day = orders_of_day.filter(subsidiary_id=subsidiary_id)
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # Crear estructura de datos para adelantos (ingresos del día)
            advances_grouped = {}
            for order in orders_of_day:
                order_cashflows_day = cashflows.filter(
                    order=order,
                    type='E',
                    transaction_date=report_date
                ).order_by('id')
                
                if order_cashflows_day.exists():
                    total_paid = sum(decimal.Decimal(cf.total) for cf in order_cashflows_day)
                    saldo = decimal.Decimal(order.total) - total_paid
                    is_paid_in_full = abs(saldo) < 0.01
                    
                    advances_grouped[order.id] = {
                        'order': order,
                        'cashflows': list(order_cashflows_day),
                        'total_advances': total_paid,
                        'saldo': saldo,
                        'is_paid_in_full': is_paid_in_full,
                        'cashflow_count': order_cashflows_day.count()
                    }
            
            # Preparar datos de saldos (cancelaciones)
            payments_cashflows = order_cashflows.filter(
                type='E',
                order_type_entry='T'
            ).exclude(
                order__register_date__gte=report_date
            )
            
            # Preparar datos de egresos
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'
            )
            
            # Calcular totales
            total_advances = sum(data['total_advances'] for data in advances_grouped.values())
            total_payments = payments_cashflows.aggregate(total=Sum('total'))['total'] or 0
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular totales por tipo de pago
            advances_efectivo = 0
            advances_yape = 0
            advances_deposito = 0
            
            for data in advances_grouped.values():
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        advances_efectivo += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        advances_yape += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'D':
                        advances_deposito += decimal.Decimal(cashflow.total)
            
            payments_efectivo = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
            payments_yape = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
            
            total_efectivo = advances_efectivo + payments_efectivo
            total_yape = advances_yape + payments_yape
            total_general = total_efectivo + total_yape
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Reporte {report_date}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill_primary = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
            header_fill_success = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            header_fill_danger = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            title_font = Font(bold=True, size=16, color="007bff")
            border = Border(
                left=Side(style='medium', color='adb5bd'),
                right=Side(style='medium', color='adb5bd'),
                top=Side(style='medium', color='adb5bd'),
                bottom=Side(style='medium', color='adb5bd')
            )
            
            # Título principal
            ws.merge_cells('A1:J1')
            ws['A1'] = f"TIENDA: {subsidiary_obj.name.upper() if subsidiary_obj else 'TODAS'} - DÍA: {datetime.strptime(report_date, '%Y-%m-%d').strftime('%d-%m-%Y')}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Sección de INGRESOS DEL DÍA
            ws['A3'] = "INGRESOS DEL DÍA"
            ws['A3'].font = header_font
            ws['A3'].fill = header_fill_primary
            ws.merge_cells('A3:J3')
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # Encabezados de ingresos
            income_headers = ['N° CPTE.', 'CLIENTE O RAZON SOCIAL', 'CANT.', 'DESCRIPCIÓN DEL PRODUCTO', 'USUARIO', 'TIPO PAGO', 'A CUENTA S/.', 'SALDO S/.', 'TOTAL S/.']
            for col, header in enumerate(income_headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_primary
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de ingresos del día
            row = 5
            for order_id, data in advances_grouped.items():
                if not data['is_paid_in_full']:  # Solo adelantos
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                            ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                            ws.cell(row=row, column=3, value=1).border = border  # Cantidad
                            # Descripción del producto
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            ws.cell(row=row, column=4, value=product_desc).border = border
                        else:
                            # Filas adicionales sin datos de orden
                            ws.cell(row=row, column=1, value="").border = border
                            ws.cell(row=row, column=2, value="").border = border
                            ws.cell(row=row, column=3, value="").border = border
                            ws.cell(row=row, column=4, value="").border = border
                        
                        # Datos del cashflow
                        ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                        
                        # Tipo de pago
                        payment_type = ""
                        if cashflow.way_to_pay == 'E':
                            payment_type = "EFECTIVO"
                        elif cashflow.way_to_pay == 'Y':
                            payment_type = "YAPE"
                        elif cashflow.way_to_pay == 'D':
                            payment_type = "DEPÓSITO"
                        ws.cell(row=row, column=6, value=payment_type).border = border
                        ws.cell(row=row, column=7, value=Decimal(cashflow.total)).border = border
                        
                        if i == 0:  # Solo en la primera fila
                            ws.cell(row=row, column=8, value=Decimal(data['saldo'])).border = border
                            ws.cell(row=row, column=9, value=Decimal(data['order'].total)).border = border
                        else:
                            ws.cell(row=row, column=8, value="").border = border
                            ws.cell(row=row, column=9, value="").border = border
                        
                row += 1
            
                # Pagos completos
                if data['is_paid_in_full']:
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                            ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                            ws.cell(row=row, column=3, value=1).border = border
                            # Descripción del producto
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            ws.cell(row=row, column=4, value=product_desc).border = border
                        else:
                            ws.cell(row=row, column=1, value="").border = border
                            ws.cell(row=row, column=2, value="").border = border
                            ws.cell(row=row, column=3, value="").border = border
                            ws.cell(row=row, column=4, value="").border = border
                        
                        ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                        
                        payment_type = ""
                        if cashflow.way_to_pay == 'E':
                            payment_type = "EFECTIVO"
                        elif cashflow.way_to_pay == 'Y':
                            payment_type = "YAPE"
                        elif cashflow.way_to_pay == 'D':
                            payment_type = "DEPÓSITO"
                        ws.cell(row=row, column=6, value=payment_type).border = border
                        ws.cell(row=row, column=7, value=decimal.Decimal(cashflow.total)).border = border
                        
                        if i == 0:
                            ws.cell(row=row, column=8, value="PAGADO").border = border
                            ws.cell(row=row, column=9, value=Decimal(data['order'].total)).border = border
                        else:
                            ws.cell(row=row, column=8, value="").border = border
                            ws.cell(row=row, column=9, value="").border = border
                        
                        row += 1
            
            # Totales de ingresos
            row += 1
            ws.cell(row=row, column=8, value="YAPE:").font = Font(bold=True)
            ws.cell(row=row, column=9, value=decimal.Decimal(advances_yape)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=8, value="EFECTIVO:").font = Font(bold=True)
            ws.cell(row=row, column=9, value=decimal.Decimal(advances_efectivo)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=8, value="TOTAL INGRESOS:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=8).fill = header_fill_primary
            ws.cell(row=row, column=9, value=decimal.Decimal(total_advances)).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=9).fill = header_fill_primary
            
            # Sección de SALDOS
            row += 3
            ws.cell(row=row, column=1, value="SALDOS")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_success
            ws.merge_cells(f'A{row}:F{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Encabezados de saldos
            row += 1
            saldos_headers = ['N° COMPROBANTE', 'FECHA', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'S/TOTAL']
            for col, header in enumerate(saldos_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_success
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de saldos
            row += 1
            for cashflow in payments_cashflows:
                ws.cell(row=row, column=1, value=f"{cashflow.order.subsidiary.serial}-{cashflow.order.correlative:03d}").border = border
                ws.cell(row=row, column=2, value=cashflow.order.register_date.strftime('%d-%m-%Y')).border = border
                ws.cell(row=row, column=3, value=cashflow.description or "PAGO TOTAL").border = border
                ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                
                payment_type = ""
                if cashflow.way_to_pay == 'E':
                    payment_type = "EFECTIVO"
                elif cashflow.way_to_pay == 'Y':
                    payment_type = "YAPE"
                elif cashflow.way_to_pay == 'D':
                    payment_type = "DEPÓSITO"
                ws.cell(row=row, column=5, value=payment_type).border = border
                ws.cell(row=row, column=6, value=decimal.Decimal(cashflow.total)).border = border
                row += 1
            
            # Totales de saldos
            row += 1
            ws.cell(row=row, column=5, value="YAPE:").font = Font(bold=True)
            ws.cell(row=row, column=6, value=decimal.Decimal(payments_yape)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=5, value="EFECTIVO:").font = Font(bold=True)
            ws.cell(row=row, column=6, value=decimal.Decimal(payments_efectivo)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=5, value="TOTAL CANCELACIONES:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=5).fill = header_fill_success
            ws.cell(row=row, column=6, value=decimal.Decimal(total_payments)).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=6).fill = header_fill_success
            
            # Sección de EGRESOS
            row += 3
            ws.cell(row=row, column=1, value="EGRESOS")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_danger
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Encabezados de egresos
            row += 1
            egresos_headers = ['NRO', 'DESCRIPCIÓN', 'TIPO EGRESO', 'USUARIO', 'MONTO']
            for col, header in enumerate(egresos_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_danger
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de egresos
            row += 1
            for i, cashflow in enumerate(expenses_cashflows, 1):
                ws.cell(row=row, column=1, value=i).border = border
                ws.cell(row=row, column=2, value=cashflow.description or '-').border = border
                
                expense_type = ""
                if cashflow.type_expense == 'V':
                    expense_type = "VARIABLE"
                elif cashflow.type_expense == 'F':
                    expense_type = "FIJO"
                elif cashflow.type_expense == 'P':
                    expense_type = "PERSONAL"
                elif cashflow.type_expense == 'O':
                    expense_type = "OTRO"
                ws.cell(row=row, column=3, value=expense_type).border = border
                ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                ws.cell(row=row, column=5, value=decimal.Decimal(cashflow.total)).border = border
                row += 1
            
            # Total de egresos
            row += 1
            ws.cell(row=row, column=4, value="TOTAL EGRESOS:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=4).fill = header_fill_danger
            ws.cell(row=row, column=5, value=decimal.Decimal(total_expenses_amount)).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=5).fill = header_fill_danger
            
            # Sección de RESUMENES
            row += 3
            ws.cell(row=row, column=1, value="RESUMENES")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_success
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Resumen de ingresos
            row += 1
            ws.cell(row=row, column=1, value="INGRESOS").font = Font(bold=True, size=12, color="007bff")
            row += 1
            ws.cell(row=row, column=1, value="INGRESOS DEL DÍA:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=decimal.Decimal(total_advances)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value="SALDOS:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=decimal.Decimal(total_payments)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value="SUBTOTAL INGRESOS:").font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_advances + total_payments)).font = Font(bold=True)
            ws.cell(row=row, column=2).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            
            # Resumen de egresos
            row += 2
            ws.cell(row=row, column=1, value="EGRESOS").font = Font(bold=True, size=12, color="dc3545")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL EGRESOS:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=decimal.Decimal(total_expenses_amount)).font = Font(bold=True)
            
            # Resumen final
            row += 3
            ws.cell(row=row, column=1, value="TOTAL EFECTIVO:").font = Font(bold=True, size=11, color="28a745")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_efectivo)).font = Font(bold=True, size=11, color="28a745")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL YAPE:").font = Font(bold=True, size=11, color="17a2b8")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_yape)).font = Font(bold=True, size=11, color="17a2b8")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL EGRESOS:").font = Font(bold=True, size=11, color="dc3545")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_expenses_amount)).font = Font(bold=True, size=11, color="dc3545")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL FINAL:").font = Font(bold=True, size=12, color="ffc107")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            ws.cell(row=row, column=2, value=decimal.Decimal(total_general - total_expenses_amount)).font = Font(bold=True, size=12, color="ffc107")
            ws.cell(row=row, column=2).fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            
            # Ajustar ancho de columnas
            column_widths = [15, 25, 8, 30, 15, 12, 12, 12, 12, 12, 12, 12, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Guardar archivo
            filename = f"reporte_ventas_gastos_{report_date}.xlsx"
            filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            wb.save(filepath)
            
            # Retornar URL del archivo
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte exportado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar el reporte: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)


@csrf_exempt
def export_sales_report_pdf(request):
    """Exportar reporte de ventas a PDF"""
    if request.method == 'POST':
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            # Obtener datos del reporte
            report_date = request.POST.get('report_date')
            subsidiary_id = request.POST.get('subsidiary')
            cash_id = request.POST.get('cash_account')
            subsidiary_obj = None
            
            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por sucursal
            if subsidiary_id and subsidiary_id != '0':
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    cash__subsidiary_id=subsidiary_id
                )
            else:
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']
            ).order_by('id')
            
            if subsidiary_id and subsidiary_id != '0':
                orders_of_day = orders_of_day.filter(subsidiary_id=subsidiary_id)
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # Crear estructura de datos para adelantos (ingresos del día)
            advances_grouped = {}
            for order in orders_of_day:
                order_cashflows_day = cashflows.filter(
                    order=order,
                    type='E',
                    transaction_date=report_date
                ).order_by('id')
                
                if order_cashflows_day.exists():
                    total_paid = sum(decimal.Decimal(cf.total) for cf in order_cashflows_day)
                    saldo = decimal.Decimal(order.total) - total_paid
                    is_paid_in_full = abs(saldo) < 0.01
                    
                    advances_grouped[order.id] = {
                        'order': order,
                        'cashflows': list(order_cashflows_day),
                        'total_advances': total_paid,
                        'saldo': saldo,
                        'is_paid_in_full': is_paid_in_full,
                        'cashflow_count': order_cashflows_day.count()
                    }
            
            # Preparar datos de saldos (cancelaciones)
            payments_cashflows = order_cashflows.filter(
                type='E',
                order_type_entry='T'
            ).exclude(
                order__register_date__gte=report_date
            )
            
            # Preparar datos de egresos
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'
            )
            
            # Calcular totales
            total_advances = sum(data['total_advances'] for data in advances_grouped.values())
            total_payments = payments_cashflows.aggregate(total=Sum('total'))['total'] or 0
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular totales por tipo de pago
            advances_efectivo = 0
            advances_yape = 0
            
            for data in advances_grouped.values():
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        advances_efectivo += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        advances_yape += decimal.Decimal(cashflow.total)
            
            payments_efectivo = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
            payments_yape = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
            
            total_efectivo = advances_efectivo + payments_efectivo
            total_yape = advances_yape + payments_yape
            total_general = total_efectivo + total_yape
            
            # Crear archivo PDF
            filename = f"reporte_ventas_gastos_{report_date}.pdf"
            filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            doc = SimpleDocTemplate(filepath, pagesize=landscape(letter), 
                                  leftMargin=0.5*inch, rightMargin=0.5*inch, 
                                  topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1,
                textColor=colors.HexColor('#007bff')
            )
            
            # Título
            subsidiary_name = subsidiary_obj.name.upper() if subsidiary_obj else 'TODAS'
            formatted_date = datetime.strptime(report_date, '%Y-%m-%d').strftime('%d-%m-%Y')
            story.append(Paragraph(f"TIENDA: {subsidiary_name} - DÍA: {formatted_date}", title_style))
            story.append(Spacer(1, 20))
            
            # Tabla de INGRESOS DEL DÍA
            income_data = [['N° CPTE.', 'CLIENTE', 'CANT.', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'A CUENTA S/.', 'SALDO S/.', 'TOTAL S/.']]
            
            # Crear estilo para párrafos en celdas
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=8,
                leading=10,
                alignment=1,  # Center alignment
                leftIndent=0,
                rightIndent=0
            )
            
            for order_id, data in advances_grouped.items():
                if not data['is_paid_in_full']:  # Solo adelantos
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            
                            payment_type = ""
                            if cashflow.way_to_pay == 'E':
                                payment_type = "EFECTIVO"
                            elif cashflow.way_to_pay == 'Y':
                                payment_type = "YAPE"
                            elif cashflow.way_to_pay == 'D':
                                payment_type = "DEPÓSITO"
                            
                            income_data.append([
                                f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}",
                                data['order'].client.full_name if data['order'].client else '-',
                                '1',
                                Paragraph(product_desc, cell_style),
                                cashflow.user.first_name or cashflow.user.username or '-',
                                payment_type,
                                f"S/. {decimal.Decimal(cashflow.total):.2f}",
                                f"S/. {Decimal(data['saldo']):.2f}",
                                f"S/. {Decimal(data['order'].total):.2f}"
                            ])
                        else:
                            # Filas adicionales sin datos de orden
                            payment_type = ""
                            if cashflow.way_to_pay == 'E':
                                payment_type = "EFECTIVO"
                            elif cashflow.way_to_pay == 'Y':
                                payment_type = "YAPE"
                            elif cashflow.way_to_pay == 'D':
                                payment_type = "DEPÓSITO"
                            
                            income_data.append([
                                '',
                                '',
                                '',
                                '',
                                cashflow.user.first_name or cashflow.user.username or '-',
                                payment_type,
                                f"S/. {decimal.Decimal(cashflow.total):.2f}",
                                '',
                                ''
                            ])
                
                # Pagos completos
                if data['is_paid_in_full']:
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            
                            payment_type = ""
                            if cashflow.way_to_pay == 'E':
                                payment_type = "EFECTIVO"
                            elif cashflow.way_to_pay == 'Y':
                                payment_type = "YAPE"
                            elif cashflow.way_to_pay == 'D':
                                payment_type = "DEPÓSITO"
                            
                            income_data.append([
                                f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}",
                                data['order'].client.full_name if data['order'].client else '-',
                                '1',
                                Paragraph(product_desc, cell_style),
                                cashflow.user.first_name or cashflow.user.username or '-',
                                payment_type,
                                f"S/. {decimal.Decimal(cashflow.total):.2f}",
                                "PAGADO",
                                f"S/. {Decimal(data['order'].total):.2f}"
                            ])
                        else:
                            payment_type = ""
                            if cashflow.way_to_pay == 'E':
                                payment_type = "EFECTIVO"
                            elif cashflow.way_to_pay == 'Y':
                                payment_type = "YAPE"
                            elif cashflow.way_to_pay == 'D':
                                payment_type = "DEPÓSITO"
                            
                            income_data.append([
                                '',
                                '',
                                '',
                                '',
                                cashflow.user.first_name or cashflow.user.username or '-',
                                payment_type,
                                f"S/. {decimal.Decimal(cashflow.total):.2f}",
                                '',
                                ''
                            ])
            
            # Agregar totales de ingresos
            income_data.append(['', '', '', '', '', 'YAPE:', '', '', f"S/. {decimal.Decimal(advances_yape):.2f}"])
            income_data.append(['', '', '', '', '', 'EFECTIVO:', '', '', f"S/. {decimal.Decimal(advances_efectivo):.2f}"])
            income_data.append(['', '', '', '', '', 'TOTAL INGRESOS:', '', '', f"S/. {decimal.Decimal(total_advances):.2f}"])
            
            income_table = Table(income_data, colWidths=[50, 120, 30, 180, 90, 70, 70, 70, 70])
            income_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, -3), (-1, -1), colors.HexColor('#007bff')),
                ('TEXTCOLOR', (0, -3), (-1, -1), colors.whitesmoke),
                ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#adb5bd'))
            ]))
            
            story.append(Paragraph("INGRESOS DEL DÍA", styles['Heading2']))
            story.append(income_table)
            story.append(Spacer(1, 20))
            
            # Tabla de SALDOS
            saldos_data = [['N° COMPROBANTE', 'FECHA', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'S/TOTAL']]
            
            for cashflow in payments_cashflows:
                payment_type = ""
                if cashflow.way_to_pay == 'E':
                    payment_type = "EFECTIVO"
                elif cashflow.way_to_pay == 'Y':
                    payment_type = "YAPE"
                elif cashflow.way_to_pay == 'D':
                    payment_type = "DEPÓSITO"
                
                saldos_data.append([
                    f"{cashflow.order.subsidiary.serial}-{cashflow.order.correlative:03d}",
                    cashflow.order.register_date.strftime('%d-%m-%Y'),
                    Paragraph(cashflow.description or "PAGO TOTAL", cell_style),
                    cashflow.user.first_name or cashflow.user.username or '-',
                    payment_type,
                    f"S/. {decimal.Decimal(cashflow.total):.2f}"
                ])
            
            # Agregar totales de saldos
            saldos_data.append(['', '', '', '', 'YAPE:', f"S/. {decimal.Decimal(payments_yape):.2f}"])
            saldos_data.append(['', '', '', '', 'EFECTIVO:', f"S/. {decimal.Decimal(payments_efectivo):.2f}"])
            saldos_data.append(['', '', '', '', 'TOTAL CANCELACIONES:', f"S/. {decimal.Decimal(total_payments):.2f}"])
            
            saldos_table = Table(saldos_data, colWidths=[120, 80, 200, 100, 80, 80])
            saldos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, -3), (-1, -1), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, -3), (-1, -1), colors.whitesmoke),
                ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#adb5bd'))
            ]))
            
            story.append(Paragraph("SALDOS", styles['Heading2']))
            story.append(saldos_table)
            story.append(Spacer(1, 20))
            
            # Tabla de EGRESOS
            egresos_data = [['NRO', 'DESCRIPCIÓN', 'TIPO EGRESO', 'USUARIO', 'MONTO']]
            
            for i, cashflow in enumerate(expenses_cashflows, 1):
                expense_type = ""
                if cashflow.type_expense == 'V':
                    expense_type = "VARIABLE"
                elif cashflow.type_expense == 'F':
                    expense_type = "FIJO"
                elif cashflow.type_expense == 'P':
                    expense_type = "PERSONAL"
                elif cashflow.type_expense == 'O':
                    expense_type = "OTRO"
                
                egresos_data.append([
                    str(i),
                    Paragraph(cashflow.description or '-', cell_style),
                    expense_type,
                    cashflow.user.first_name or cashflow.user.username or '-',
                    f"S/. {decimal.Decimal(cashflow.total):.2f}"
                ])
            
            # Agregar total de egresos
            egresos_data.append(['', '', '', 'TOTAL EGRESOS:', f"S/. {decimal.Decimal(total_expenses_amount):.2f}"])
            
            egresos_table = Table(egresos_data, colWidths=[50, 250, 100, 100, 80])
            egresos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#dc3545')),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#adb5bd'))
            ]))
            
            story.append(Paragraph("EGRESOS", styles['Heading2']))
            story.append(egresos_table)
            story.append(Spacer(1, 20))
            
            # Resumen final
            summary_data = [
                ['RESUMENES'],
                ['INGRESOS', ''],
                ['INGRESOS DEL DÍA:', f"S/. {decimal.Decimal(total_advances):.2f}"],
                ['SALDOS:', f"S/. {decimal.Decimal(total_payments):.2f}"],
                ['SUBTOTAL INGRESOS:', f"S/. {decimal.Decimal(total_advances + total_payments):.2f}"],
                ['', ''],
                ['EGRESOS', ''],
                ['TOTAL EGRESOS:', f"S/. {decimal.Decimal(total_expenses_amount):.2f}"],
                ['', ''],
                ['TOTAL EFECTIVO:', f"S/. {decimal.Decimal(total_efectivo):.2f}"],
                ['TOTAL YAPE:', f"S/. {decimal.Decimal(total_yape):.2f}"],
                ['TOTAL EGRESOS:', f"S/. {decimal.Decimal(total_expenses_amount):.2f}"],
                ['TOTAL FINAL:', f"S/. {decimal.Decimal(total_general - total_expenses_amount):.2f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[120, 80])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#007bff')),
                ('TEXTCOLOR', (0, 1), (-1, 1), colors.whitesmoke),
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#dc3545')),
                ('TEXTCOLOR', (0, 6), (-1, 6), colors.whitesmoke),
                ('FONTNAME', (0, 6), (-1, 6), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ffc107')),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#adb5bd'))
            ]))
            
            story.append(Paragraph("RESUMENES", styles['Heading2']))
            story.append(summary_table)
            
            # Generar PDF
            doc.build(story)
            
            # Retornar URL del archivo
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte PDF generado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al generar el PDF: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


def sales_report_by_user(request):
    """Vista principal del reporte de ventas y gastos por usuario"""
    if request.method == 'GET':
        from apps.users.models import CustomUser
        
        # Obtener todos los usuarios que tienen acceso al sistema
        users_set = CustomUser.objects.filter(
            has_access_system=True,
            is_active=True,
            is_staff=False
        ).order_by('first_name', 'last_name')
        
        # Fecha actual para el filtro
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        return render(request, 'accounting/sales_report_by_user.html', {
            'users_set': users_set,
            'date_now': date_now,
            'current_user': request.user,
        })
    elif request.method == 'POST':
        try:
            from apps.users.models import CustomUser
            
            # Obtener parámetros del filtro
            report_date = request.POST.get('report_date')
            user_id = int(request.POST.get('user'))
            user_obj = None

            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por usuario
            if user_id and user_id != '0':
                user_obj = CustomUser.objects.get(id=int(user_id))
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    user_id=user_id
                )
            else:
                # Si no hay usuario específico, mostrar todos los cashflows del día
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            # Excluir órdenes anuladas (status='A')
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Calcular totales desde accounting_cashflow
            # total_sales: Suma total de todas las órdenes relacionadas con cashflows del día
            total_sales = sum(cf.order.total for cf in order_cashflows if cf.order) or 0
            
            # total_cash_advance: Suma de adelantos (tipo 'A') desde accounting_cashflow
            total_cash_advance = order_cashflows.filter(
                type='E',  # Entrada
                order_type_entry='A'  # Adelanto
            ).aggregate(total=Sum('total'))['total'] or 0
            
            # total_paid: Suma de pagos totales (tipo 'T') desde accounting_cashflow
            total_paid = order_cashflows.filter(
                type='E',  # Entrada
                order_type_entry='T'  # Pago Total
            ).aggregate(total=Sum('total'))['total'] or 0
            
            # total_balance: Saldo pendiente (ventas totales - adelantos totales - pagos totales)
            total_balance = total_sales - total_cash_advance - total_paid
            
            # Calcular totales de gastos
            total_income = cashflows.filter(type='E').aggregate(total=Sum('total'))['total'] or 0
            total_expenses = cashflows.filter(type='S').aggregate(total=Sum('total'))['total'] or 0
            net_expenses = total_expenses - total_income
            
            # Calcular subtotales por tipo de gasto según TYPE_EXPENSE
            total_variable_expenses = cashflows.filter(type='S', type_expense='V').aggregate(total=Sum('total'))['total'] or 0
            total_fixed_expenses = cashflows.filter(type='S', type_expense='F').aggregate(total=Sum('total'))['total'] or 0
            total_personal_expenses = cashflows.filter(type='S', type_expense='P').aggregate(total=Sum('total'))['total'] or 0
            total_other_expenses = cashflows.filter(type='S', type_expense='O').aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular ingreso real: A Cuenta + Total Pagado
            real_income = total_cash_advance + total_paid
            
            # Calcular caja final
            # final_cash: Caja final = Ingreso real - Gastos netos
            # Donde: Gastos netos = Gastos totales - Ingresos por otros conceptos
            final_cash = real_income - net_expenses

            # NUEVA LÓGICA: 
            # 1. ADELANTOS = Todas las órdenes del día (con sus cashflows)
            # 2. SALDOS = Solo cashflows de cancelación (order_type_entry='T')
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']  # Pendientes y completadas
            ).order_by('id')
            
            # Si hay filtro por usuario, incluir órdenes que:
            # 1. Fueron creadas por el usuario, O
            # 2. Tienen cashflows del día realizados por el usuario
            if user_id and user_id != '0':
                # Obtener IDs de órdenes que tienen cashflows del usuario del día
                orders_with_user_cashflows = order_cashflows.filter(
                    user_id=user_id
                ).values_list('order_id', flat=True).distinct()
                
                # Filtrar órdenes: creadas por el usuario O con cashflows del usuario
                orders_of_day = orders_of_day.filter(
                    Q(user_id=user_id) | Q(id__in=orders_with_user_cashflows)
                )
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # ========================================
            # DICT 1: ADELANTOS DEL USUARIO
            # ========================================
            adelantos_usuario = {}
            
            for order in orders_of_day:
                order_advances = cashflows.filter(
                    order=order,
                    type='E',  # Solo entradas
                    order_type_entry='A',  # Solo adelantos
                    transaction_date=report_date,  # Solo del día del reporte
                    user_id=user_id  # Solo hechos por el usuario
                ).order_by('id')
                
                if order_advances.exists():
                    total_advances = sum(float(cf.total) for cf in order_advances)
                    saldo = float(order.total) - total_advances
                    is_paid_in_full = abs(saldo) < 0.01
                    
                    adelantos_usuario[f"advance_{order.id}"] = {
                        'tipo': 'adelanto',
                        'order': order,
                        'cashflows': list(order_advances),
                        'total_amount': total_advances,
                        'saldo': saldo,
                        'is_paid_in_full': is_paid_in_full,
                        'cashflow_count': order_advances.count()
                    }
            
            # ========================================
            # DICT 2: PAGOS TOTALES DE ÓRDENES DEL USUARIO
            # ========================================
            pagos_totales_usuario = {}
            
            for order in orders_of_day:
                # Solo procesar órdenes creadas por el usuario
                if order.user_id == int(user_id):
                    order_payments = cashflows.filter(
                        order=order,
                        type='E',  # Solo entradas
                        order_type_entry='T',  # Solo pagos totales
                        transaction_date=report_date,  # Solo del día del reporte
                        user_id=user_id  # Solo hechos por el usuario
                    ).order_by('id')
                    
                    if order_payments.exists():
                        total_payments = sum(float(cf.total) for cf in order_payments)
                        
                        pagos_totales_usuario[f"payment_{order.id}"] = {
                            'tipo': 'pago_total_usuario',
                            'order': order,
                            'cashflows': list(order_payments),
                            'total_amount': total_payments,
                            'cashflow_count': order_payments.count()
                        }
            
            # ========================================
            # DICT 3: PAGOS TOTALES DE ÓRDENES NO DEL USUARIO
            # ========================================
            pagos_totales_otros = {}
            
            other_orders_payments = cashflows.filter(
                type='E',  # Solo entradas
                order_type_entry='T',  # Solo pagos totales
                transaction_date=report_date,  # Solo del día del reporte
                user_id=user_id,  # Hechos por el usuario
            ).exclude(
                order__user_id=user_id  # Excluir órdenes creadas por el usuario
            ).order_by('order_id', 'id')
            
            for cashflow in other_orders_payments:
                pagos_totales_otros[f"cancellation_{cashflow.id}"] = {
                    'tipo': 'pago_total_otros',
                    'order': cashflow.order,
                    'cashflows': [cashflow],
                    'total_amount': float(cashflow.total),
                    'cashflow_count': 1
                }
            
            # ========================================
            # COMBINAR LOS TRES DICT EN INGRESOS DEL DÍA
            # ========================================
            ingresos_del_dia = {}
            
            # 1. Agregar adelantos
            ingresos_del_dia.update(adelantos_usuario)
            
            # 2. Agregar separador
            ingresos_del_dia['separador'] = {'tipo': 'separador'}
            
            # 3. Agregar pagos totales del usuario
            ingresos_del_dia.update(pagos_totales_usuario)
            
            # 4. Agregar pagos totales de otros
            ingresos_del_dia.update(pagos_totales_otros)
            
            # ========================================
            # DICT 2: PAGOS DE FECHAS ANTERIORES
            # ========================================
            pagos_fechas_anteriores = {}
            
            # Pagos totales de fechas anteriores hechas por el usuario

            previous_payments = cashflows.filter(
                type='E',  # Solo entradas
                order_type_entry='T',  # Solo pagos totales
                user_id=user_id,  # Hechos por el usuario
                transaction_date=report_date,  # Pagados en la fecha del reporte
                order__register_date__lt=report_date  # De órdenes de fechas anteriores
            ).order_by('order_id', 'id')
            
            for cashflow in previous_payments:
                pagos_fechas_anteriores[f"previous_{cashflow.id}"] = {
                    'order': cashflow.order,
                    'cashflows': [cashflow],
                    'total_amount': float(cashflow.total),
                    'transaction_date': cashflow.order.register_date,
                    'cashflow_count': 1
                }
            
            # ========================================
            # EGRESOS (mantener como estaba)
            # ========================================
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'  # Solo salidas (gastos)
            )

            # ========================================
            # CÁLCULO DE TOTALES
            # ========================================
            
            # Total de apertura de caja (tipo 'A')
            total_apertura = cashflows.filter(type='A').aggregate(total=Sum('total'))['total'] or 0
            
            # Totales de ingresos del día
            total_ingresos_dia = 0
            ingresos_efectivo = 0
            ingresos_yape = 0
            ingresos_deposito = 0
            
            for key, data in ingresos_del_dia.items():
                if data['tipo'] != 'separador':
                    total_ingresos_dia += data['total_amount']
                    for cashflow in data['cashflows']:
                        if cashflow.way_to_pay == 'E':
                            ingresos_efectivo += decimal.Decimal(cashflow.total)
                        elif cashflow.way_to_pay == 'Y':
                            ingresos_yape += decimal.Decimal(cashflow.total)
                        elif cashflow.way_to_pay == 'D':
                            ingresos_deposito += decimal.Decimal(cashflow.total)
            
            # Totales de pagos de fechas anteriores
            total_pagos_anteriores = 0
            pagos_anteriores_efectivo = 0
            pagos_anteriores_yape = 0
            pagos_anteriores_deposito = 0
            
            for key, data in pagos_fechas_anteriores.items():
                total_pagos_anteriores += data['total_amount']
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        pagos_anteriores_efectivo += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        pagos_anteriores_yape += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'D':
                        pagos_anteriores_deposito += decimal.Decimal(cashflow.total)
            
            # Totales de egresos
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            # Totales generales
            total_efectivo = ingresos_efectivo + pagos_anteriores_efectivo
            total_yape = ingresos_yape + pagos_anteriores_yape
            total_deposito = ingresos_deposito + pagos_anteriores_deposito
            total_general = total_efectivo + total_yape + total_deposito + total_apertura

            context = {
                'report_date': datetime.strptime(report_date, "%Y-%m-%d").strftime("%d-%m-%Y"),
                'ingresos_del_dia': ingresos_del_dia,  # Nuevo dict con ingresos del día
                'pagos_fechas_anteriores': pagos_fechas_anteriores,  # Nuevo dict con pagos anteriores
                'orders_of_day': orders_of_day,  # Todas las órdenes del día
                'expenses_cashflows': expenses_cashflows.order_by('id'),
                'total_ingresos_dia': total_ingresos_dia,  # Total ingresos del día
                'total_pagos_anteriores': total_pagos_anteriores,  # Total pagos anteriores
                'total_expenses_amount': total_expenses_amount,
                'total_apertura': total_apertura,  # Total apertura de caja
                'ingresos_efectivo': ingresos_efectivo,
                'ingresos_yape': ingresos_yape,
                'ingresos_deposito': ingresos_deposito,
                'pagos_anteriores_efectivo': pagos_anteriores_efectivo,
                'pagos_anteriores_yape': pagos_anteriores_yape,
                'pagos_anteriores_deposito': pagos_anteriores_deposito,
                'total_efectivo': total_efectivo,
                'total_yape': total_yape,
                'total_deposito': total_deposito,
                'total_general': total_general,
                'user': user_obj,
            }
            
            tpl = loader.get_template('accounting/sales_report_by_user_grid.html')
            
            return JsonResponse({
                'success': True,
                'grid': tpl.render(context, request),
                'summary': {
                    'total_sales': float(total_sales),
                    'total_cash_advance': float(total_cash_advance),
                    'total_paid': float(total_paid),
                    'total_balance': float(total_balance),
                    'real_income': float(real_income),
                    'total_income': float(total_income),
                    'total_expenses': float(total_expenses),
                    'net_expenses': float(net_expenses),
                    'final_cash': float(final_cash),
                    'total_variable_expenses': float(total_variable_expenses),
                    'total_fixed_expenses': float(total_fixed_expenses),
                    'total_personal_expenses': float(total_personal_expenses),
                    'total_other_expenses': float(total_other_expenses),
                }
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            print(f"ERROR en sales_report_by_user: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'message': f'Error al generar el reporte: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)


@csrf_exempt
def export_sales_report_by_user_excel(request):
    """Exportar reporte de ventas por usuario a Excel"""
    if request.method == 'POST':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from decimal import Decimal
            
            # Obtener datos del reporte
            report_date = request.POST.get('report_date')
            user_id = request.POST.get('user')
            user_obj = None
            
            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por usuario
            if user_id and user_id != '0':
                from apps.users.models import CustomUser
                user_obj = CustomUser.objects.get(id=int(user_id))
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    user_id=user_id
                )
            else:
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']
            ).order_by('id')
            
            if user_id and user_id != '0':
                # Obtener IDs de órdenes que tienen cashflows del usuario del día
                orders_with_user_cashflows = order_cashflows.filter(
                    user_id=user_id
                ).values_list('order_id', flat=True).distinct()
                
                # Filtrar órdenes: creadas por el usuario O con cashflows del usuario
                orders_of_day = orders_of_day.filter(
                    Q(user_id=user_id) | Q(id__in=orders_with_user_cashflows)
                )
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # Crear estructura de datos para adelantos (ingresos del día)
            advances_grouped = {}
            for order in orders_of_day:
                order_cashflows_day = cashflows.filter(
                    order=order,
                    type='E',
                    transaction_date=report_date
                ).order_by('id')
                
                if order_cashflows_day.exists():
                    total_paid = sum(decimal.Decimal(cf.total) for cf in order_cashflows_day)
                    saldo = decimal.Decimal(order.total) - total_paid
                    is_paid_in_full = abs(saldo) < 0.01
                    
                    advances_grouped[order.id] = {
                        'order': order,
                        'cashflows': list(order_cashflows_day),
                        'total_advances': total_paid,
                        'saldo': saldo,
                        'is_paid_in_full': is_paid_in_full,
                        'cashflow_count': order_cashflows_day.count()
                    }
            
            # Preparar datos de saldos (cancelaciones)
            payments_cashflows = order_cashflows.filter(
                type='E',
                order_type_entry='T'
            ).exclude(
                order__register_date__gte=report_date
            )
            
            # Preparar datos de egresos
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'
            )
            
            # Calcular totales
            total_advances = sum(data['total_advances'] for data in advances_grouped.values())
            total_payments = payments_cashflows.aggregate(total=Sum('total'))['total'] or 0
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular totales por tipo de pago
            advances_efectivo = 0
            advances_yape = 0
            advances_deposito = 0
            
            for data in advances_grouped.values():
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        advances_efectivo += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        advances_yape += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'D':
                        advances_deposito += decimal.Decimal(cashflow.total)
            
            payments_efectivo = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
            payments_yape = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
            
            total_efectivo = advances_efectivo + payments_efectivo
            total_yape = advances_yape + payments_yape
            total_general = total_efectivo + total_yape
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Reporte {report_date}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill_primary = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
            header_fill_success = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            header_fill_danger = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            title_font = Font(bold=True, size=16, color="007bff")
            border = Border(
                left=Side(style='medium', color='adb5bd'),
                right=Side(style='medium', color='adb5bd'),
                top=Side(style='medium', color='adb5bd'),
                bottom=Side(style='medium', color='adb5bd')
            )
            
            # Título principal
            ws.merge_cells('A1:J1')
            user_name = f"{user_obj.first_name} {user_obj.last_name}".strip() if user_obj else "TODOS"
            ws['A1'] = f"USUARIO: {user_name.upper()} - DÍA: {datetime.strptime(report_date, '%Y-%m-%d').strftime('%d-%m-%Y')}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Sección de INGRESOS DEL DÍA
            ws['A3'] = "INGRESOS DEL DÍA"
            ws['A3'].font = header_font
            ws['A3'].fill = header_fill_primary
            ws.merge_cells('A3:J3')
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # Encabezados de ingresos
            income_headers = ['N° CPTE.', 'CLIENTE O RAZON SOCIAL', 'CANT.', 'DESCRIPCIÓN DEL PRODUCTO', 'USUARIO', 'TIPO PAGO', 'A CUENTA S/.', 'SALDO S/.', 'TOTAL S/.']
            for col, header in enumerate(income_headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_primary
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de ingresos del día
            row = 5
            for order_id, data in advances_grouped.items():
                if not data['is_paid_in_full']:  # Solo adelantos
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                            ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                            ws.cell(row=row, column=3, value=1).border = border  # Cantidad
                            # Descripción del producto
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            ws.cell(row=row, column=4, value=product_desc).border = border
                        else:
                            # Filas adicionales sin datos de orden
                            ws.cell(row=row, column=1, value="").border = border
                            ws.cell(row=row, column=2, value="").border = border
                            ws.cell(row=row, column=3, value="").border = border
                            ws.cell(row=row, column=4, value="").border = border
                        
                        # Datos del cashflow
                        ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                        
                        # Tipo de pago
                        payment_type = ""
                        if cashflow.way_to_pay == 'E':
                            payment_type = "EFECTIVO"
                        elif cashflow.way_to_pay == 'Y':
                            payment_type = "YAPE"
                        elif cashflow.way_to_pay == 'D':
                            payment_type = "DEPÓSITO"
                        ws.cell(row=row, column=6, value=payment_type).border = border
                        ws.cell(row=row, column=7, value=Decimal(cashflow.total)).border = border
                        
                        if i == 0:  # Solo en la primera fila
                            ws.cell(row=row, column=8, value=Decimal(data['saldo'])).border = border
                            ws.cell(row=row, column=9, value=Decimal(data['order'].total)).border = border
                        else:
                            ws.cell(row=row, column=8, value="").border = border
                            ws.cell(row=row, column=9, value="").border = border
                        
                row += 1
                
                # Pagos completos
                if data['is_paid_in_full']:
                    for i, cashflow in enumerate(data['cashflows']):
                        if i == 0:  # Primera fila con datos de la orden
                            ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                            ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                            ws.cell(row=row, column=3, value=1).border = border
                            # Descripción del producto
                            product_desc = ""
                            if data['order'].orderdetail_set.exists():
                                product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                            else:
                                product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                            ws.cell(row=row, column=4, value=product_desc).border = border
                        else:
                            ws.cell(row=row, column=1, value="").border = border
                            ws.cell(row=row, column=2, value="").border = border
                            ws.cell(row=row, column=3, value="").border = border
                            ws.cell(row=row, column=4, value="").border = border
                        
                        ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                        
                        payment_type = ""
                        if cashflow.way_to_pay == 'E':
                            payment_type = "EFECTIVO"
                        elif cashflow.way_to_pay == 'Y':
                            payment_type = "YAPE"
                        elif cashflow.way_to_pay == 'D':
                            payment_type = "DEPÓSITO"
                        ws.cell(row=row, column=6, value=payment_type).border = border
                        ws.cell(row=row, column=7, value=decimal.Decimal(cashflow.total)).border = border
                        
                        if i == 0:
                            ws.cell(row=row, column=8, value="PAGADO").border = border
                            ws.cell(row=row, column=9, value=Decimal(data['order'].total)).border = border
                        else:
                            ws.cell(row=row, column=8, value="").border = border
                            ws.cell(row=row, column=9, value="").border = border
                        
                        row += 1
            
            # Totales de ingresos
            row += 1
            ws.cell(row=row, column=8, value="YAPE:").font = Font(bold=True)
            ws.cell(row=row, column=9, value=Decimal(advances_yape)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=8, value="EFECTIVO:").font = Font(bold=True)
            ws.cell(row=row, column=9, value=Decimal(advances_efectivo)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=8, value="TOTAL INGRESOS:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=8).fill = header_fill_primary
            ws.cell(row=row, column=9, value=Decimal(total_advances)).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=9).fill = header_fill_primary
            
            # Sección de SALDOS (solo si hay datos)
            if payments_cashflows.exists():
                row += 3  # Espacio entre secciones
                
                # Título de saldos
                ws.cell(row=row, column=1, value="SALDOS")
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=1).fill = header_fill_success
                ws.merge_cells(f'A{row}:F{row}')
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                
                # Encabezados de saldos
                row += 1
                payments_headers = ['N° COMPROBANTE', 'FECHA', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'S/TOTAL']
                for col, header in enumerate(payments_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill_success
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Datos de saldos
                row += 1
                for cashflow in payments_cashflows:
                    ws.cell(row=row, column=1, value=f"{cashflow.order.subsidiary.serial}-{cashflow.order.correlative:03d}").border = border
                    ws.cell(row=row, column=2, value=cashflow.order.register_date.strftime('%d-%m-%Y')).border = border
                    ws.cell(row=row, column=3, value=cashflow.description or "PAGO TOTAL").border = border
                    ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                    
                    # Tipo de pago
                    payment_type = ""
                    if cashflow.way_to_pay == 'E':
                        payment_type = "EFECTIVO"
                    elif cashflow.way_to_pay == 'Y':
                        payment_type = "YAPE"
                    elif cashflow.way_to_pay == 'D':
                        payment_type = "DEPÓSITO"
                    ws.cell(row=row, column=5, value=payment_type).border = border
                    ws.cell(row=row, column=6, value=Decimal(cashflow.total)).border = border
                    row += 1
                
                # Totales de saldos
                payments_efectivo_section = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
                payments_yape_section = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
                
                row += 1
                ws.cell(row=row, column=5, value="YAPE:").font = Font(bold=True)
                ws.cell(row=row, column=6, value=Decimal(payments_yape_section)).font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=5, value="EFECTIVO:").font = Font(bold=True)
                ws.cell(row=row, column=6, value=Decimal(payments_efectivo_section)).font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=5, value="TOTAL CANCELACIONES:").font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=5).fill = header_fill_success
                ws.cell(row=row, column=6, value=Decimal(total_payments)).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=6).fill = header_fill_success
            
            # Sección de EGRESOS (solo si hay datos)
            if expenses_cashflows.exists():
                row += 3  # Espacio entre secciones
                
                # Título de egresos
                ws.cell(row=row, column=1, value="EGRESOS")
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=1).fill = header_fill_danger
                ws.merge_cells(f'A{row}:E{row}')
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                
                # Encabezados de egresos
                row += 1
                expenses_headers = ['NRO', 'DESCRIPCIÓN', 'TIPO EGRESO', 'USUARIO', 'MONTO']
                for col, header in enumerate(expenses_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill_danger
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Datos de egresos
                row += 1
                for i, cashflow in enumerate(expenses_cashflows, 1):
                    ws.cell(row=row, column=1, value=i).border = border
                    ws.cell(row=row, column=2, value=cashflow.description or '-').border = border
                    
                    # Tipo de egreso
                    expense_type = ""
                    if cashflow.type_expense == 'V':
                        expense_type = "VARIABLE"
                    elif cashflow.type_expense == 'F':
                        expense_type = "FIJO"
                    elif cashflow.type_expense == 'P':
                        expense_type = "PERSONAL"
                    elif cashflow.type_expense == 'O':
                        expense_type = "OTRO"
                    ws.cell(row=row, column=3, value=expense_type).border = border
                    ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                    ws.cell(row=row, column=5, value=Decimal(cashflow.total)).border = border
                    row += 1
                
                # Total de egresos
                row += 1
                ws.cell(row=row, column=4, value="TOTAL EGRESOS:").font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=4).fill = header_fill_danger
                ws.cell(row=row, column=5, value=Decimal(total_expenses_amount)).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=5).fill = header_fill_danger
            
            # Sección de RESUMENES
            row += 3  # Espacio entre secciones
            
            # Título de resúmenes
            ws.cell(row=row, column=1, value="RESUMENES")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = PatternFill(start_color="6f42c1", end_color="6f42c1", fill_type="solid")  # Púrpura
            ws.merge_cells(f'A{row}:B{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Calcular totales para resúmenes
            payments_efectivo_total = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
            payments_yape_total = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
            total_general = advances_efectivo + advances_yape + advances_deposito + payments_efectivo_total + payments_yape_total
            
            # Datos de resúmenes
            summary_data = [
                ['CONCEPTO', 'MONTO'],
                ['INGRESOS DEL DÍA:', Decimal(total_advances)],
                ['SALDOS:', Decimal(total_payments)],
                ['SUBTOTAL INGRESOS:', Decimal(total_advances + total_payments)],
                ['TOTAL EGRESOS:', Decimal(total_expenses_amount)],
                ['TOTAL EFECTIVO:', Decimal(advances_efectivo + payments_efectivo_total)],
                ['TOTAL YAPE:', Decimal(advances_yape + payments_yape_total)],
                ['TOTAL FINAL:', Decimal(total_general - total_expenses_amount)]
            ]
            
            row += 1
            for i, (concepto, monto) in enumerate(summary_data):
                if i == 0:  # Encabezado
                    ws.cell(row=row, column=1, value=concepto).font = header_font
                    ws.cell(row=row, column=1).fill = PatternFill(start_color="6f42c1", end_color="6f42c1", fill_type="solid")
                    ws.cell(row=row, column=2, value=monto).font = header_font
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="6f42c1", end_color="6f42c1", fill_type="solid")
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
                else:
                    ws.cell(row=row, column=1, value=concepto).font = Font(bold=True)
                    ws.cell(row=row, column=2, value=monto).font = Font(bold=True)
                    ws.cell(row=row, column=1).border = border
                    ws.cell(row=row, column=2).border = border
                row += 1
            
            # Ajustar ancho de columnas
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Guardar archivo
            filename = f"reporte_ventas_usuario_{report_date}.xlsx"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            wb.save(file_path)
            
            # URL del archivo
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte Excel generado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al generar el Excel: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


@csrf_exempt
def export_sales_report_by_user_pdf(request):
    """Exportar reporte de ventas por usuario a PDF"""
    if request.method == 'POST':
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.pdfgen import canvas
            from io import BytesIO
            
            # Obtener datos del reporte
            report_date = request.POST.get('report_date')
            user_id = request.POST.get('user')
            user_obj = None
            
            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por usuario
            if user_id and user_id != '0':
                from apps.users.models import CustomUser
                user_obj = CustomUser.objects.get(id=int(user_id))
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    user_id=user_id
                )
            else:
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']
            ).order_by('id')
            
            if user_id and user_id != '0':
                # Obtener IDs de órdenes que tienen cashflows del usuario del día
                orders_with_user_cashflows = order_cashflows.filter(
                    user_id=user_id
                ).values_list('order_id', flat=True).distinct()
                
                # Filtrar órdenes: creadas por el usuario O con cashflows del usuario
                orders_of_day = orders_of_day.filter(
                    Q(user_id=user_id) | Q(id__in=orders_with_user_cashflows)
                )
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # Crear estructura de datos para adelantos (ingresos del día)
            advances_grouped = {}
            for order in orders_of_day:
                order_cashflows_day = cashflows.filter(
                    order=order,
                    type='E',
                    transaction_date=report_date
                ).order_by('id')
                
                if order_cashflows_day.exists():
                    total_paid = sum(decimal.Decimal(cf.total) for cf in order_cashflows_day)
                    saldo = decimal.Decimal(order.total) - total_paid
                    is_paid_in_full = abs(saldo) < 0.01
                    
                    advances_grouped[order.id] = {
                        'order': order,
                        'cashflows': list(order_cashflows_day),
                        'total_advances': total_paid,
                        'saldo': saldo,
                        'is_paid_in_full': is_paid_in_full,
                        'cashflow_count': order_cashflows_day.count()
                    }
            
            # Preparar datos de saldos (cancelaciones)
            payments_cashflows = order_cashflows.filter(
                type='E',
                order_type_entry='T'
            ).exclude(
                order__register_date__gte=report_date
            )
            
            # Preparar datos de egresos
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'
            )
            
            # Calcular totales
            total_advances = sum(data['total_advances'] for data in advances_grouped.values())
            total_payments = payments_cashflows.aggregate(total=Sum('total'))['total'] or 0
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular totales por tipo de pago para adelantos (ingresos del día)
            advances_efectivo = 0
            advances_yape = 0
            advances_deposito = 0
            
            for data in advances_grouped.values():
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        advances_efectivo += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        advances_yape += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'D':
                        advances_deposito += decimal.Decimal(cashflow.total)
            
            # Crear PDF
            filename = f"reporte_ventas_usuario_{report_date}.pdf"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            
            user_name = f"{user_obj.first_name} {user_obj.last_name}".strip() if user_obj else "TODOS"
            title = Paragraph(f"USUARIO: {user_name.upper()} - DÍA: {datetime.strptime(report_date, '%Y-%m-%d').strftime('%d-%m-%Y')}", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # Sección de INGRESOS DEL DÍA
            story.append(Paragraph("INGRESOS DEL DÍA", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Crear tabla de ingresos
            income_data = [['N° CPTE.', 'CLIENTE', 'CANT.', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'A CUENTA S/.', 'SALDO S/.', 'TOTAL S/.']]
            
            for order_id, data in advances_grouped.items():
                if not data['is_paid_in_full']:  # Solo adelantos
                    for i, cashflow in enumerate(data['cashflows']):
                        row_data = []
                        if i == 0:  # Primera fila con datos de la orden
                            row_data = [
                                f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}",
                                data['order'].client.full_name if data['order'].client else '-',
                                '1',
                                data['order'].observation or "ORDEN DE SERVICIO",
                                cashflow.user.first_name or cashflow.user.username or '-',
                                'EFECTIVO' if cashflow.way_to_pay == 'E' else 'YAPE' if cashflow.way_to_pay == 'Y' else 'DEPÓSITO',
                                f"S/ {Decimal(cashflow.total):.2f}",
                                f"S/ {Decimal(data['saldo']):.2f}",
                                f"S/ {Decimal(data['order'].total):.2f}"
                            ]
                        else:
                            row_data = ['', '', '', '', cashflow.user.first_name or cashflow.user.username or '-', 
                                      'EFECTIVO' if cashflow.way_to_pay == 'E' else 'YAPE' if cashflow.way_to_pay == 'Y' else 'DEPÓSITO',
                                      f"S/ {Decimal(cashflow.total):.2f}", '', '']
                        income_data.append(row_data)
                
                # Pagos completos
                if data['is_paid_in_full']:
                    for i, cashflow in enumerate(data['cashflows']):
                        row_data = []
                        if i == 0:  # Primera fila con datos de la orden
                            row_data = [
                                f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}",
                                data['order'].client.full_name if data['order'].client else '-',
                                '1',
                                data['order'].observation or "ORDEN DE SERVICIO",
                                cashflow.user.first_name or cashflow.user.username or '-',
                                'EFECTIVO' if cashflow.way_to_pay == 'E' else 'YAPE' if cashflow.way_to_pay == 'Y' else 'DEPÓSITO',
                                f"S/ {Decimal(cashflow.total):.2f}",
                                "PAGADO",
                                f"S/ {data['order'].total}"
                            ]
                        else:
                            row_data = ['', '', '', '', cashflow.user.first_name or cashflow.user.username or '-', 
                                      'EFECTIVO' if cashflow.way_to_pay == 'E' else 'YAPE' if cashflow.way_to_pay == 'Y' else 'DEPÓSITO',
                                      f"S/ {Decimal(cashflow.total):.2f}", '', '']
                        income_data.append(row_data)
            
            # Agregar totales
            income_data.append(['', '', '', '', '', '', '', 'YAPE:', f"S/ {Decimal(advances_yape):.2f}"])
            income_data.append(['', '', '', '', '', '', '', 'EFECTIVO:', f"S/ {Decimal(advances_efectivo):.2f}"])
            income_data.append(['', '', '', '', '', '', '', 'TOTAL INGRESOS:', f"S/ {Decimal(total_advances):.2f}"])
            
            # Crear tabla
            income_table = Table(income_data)
            income_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
            ]))
            
            story.append(income_table)
            story.append(Spacer(1, 20))
            
            # Sección de SALDOS (solo si hay datos)
            if payments_cashflows.exists():
                story.append(Paragraph("SALDOS", styles['Heading2']))
                story.append(Spacer(1, 12))
                
                # Crear tabla de saldos
                payments_data = [['N° COMPROBANTE', 'FECHA', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'S/TOTAL']]
                
                for cashflow in payments_cashflows:
                    payments_data.append([
                        f"{cashflow.order.subsidiary.serial}-{cashflow.order.correlative:03d}",
                        cashflow.order.register_date.strftime('%d-%m-%Y'),
                        cashflow.description or "PAGO TOTAL",
                        cashflow.user.first_name or cashflow.user.username or '-',
                        'EFECTIVO' if cashflow.way_to_pay == 'E' else 'YAPE' if cashflow.way_to_pay == 'Y' else 'DEPÓSITO',
                        f"S/ {Decimal(cashflow.total):.2f}"
                    ])
                
                # Agregar totales de saldos
                payments_efectivo_section = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
                payments_yape_section = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
                
                payments_data.append(['', '', '', '', 'YAPE:', f"S/ {Decimal(payments_yape_section):.2f}"])
                payments_data.append(['', '', '', '', 'EFECTIVO:', f"S/ {Decimal(payments_efectivo_section):.2f}"])
                payments_data.append(['', '', '', '', 'TOTAL CANCELACIONES:', f"S/ {Decimal(total_payments):.2f}"])
                
                # Crear tabla de saldos
                payments_table = Table(payments_data)
                payments_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.green),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 6),
                ]))
                
                story.append(payments_table)
                story.append(Spacer(1, 20))
            
            # Sección de EGRESOS (solo si hay datos)
            if expenses_cashflows.exists():
                story.append(Paragraph("EGRESOS", styles['Heading2']))
                story.append(Spacer(1, 12))
                
                # Crear tabla de egresos
                expenses_data = [['NRO', 'DESCRIPCIÓN', 'TIPO EGRESO', 'USUARIO', 'MONTO']]
                
                for i, cashflow in enumerate(expenses_cashflows, 1):
                    expense_type = 'VARIABLE' if cashflow.type_expense == 'V' else 'FIJO' if cashflow.type_expense == 'F' else 'PERSONAL' if cashflow.type_expense == 'P' else 'OTRO'
                    
                    expenses_data.append([
                        str(i),
                        cashflow.description or '-',
                        expense_type,
                        cashflow.user.first_name or cashflow.user.username or '-',
                        f"S/ {Decimal(cashflow.total):.2f}"
                    ])
                
                # Agregar total de egresos
                expenses_data.append(['', '', '', 'TOTAL EGRESOS:', f"S/ {Decimal(total_expenses_amount):.2f}"])
                
                # Crear tabla de egresos
                expenses_table = Table(expenses_data)
                expenses_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 6),
                ]))
                
                story.append(expenses_table)
                story.append(Spacer(1, 20))
            
            # Sección de RESUMENES
            story.append(Paragraph("RESUMENES", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Calcular totales de pagos para resúmenes
            payments_efectivo_total = payments_cashflows.filter(way_to_pay='E').aggregate(total=Sum('total'))['total'] or 0
            payments_yape_total = payments_cashflows.filter(way_to_pay='Y').aggregate(total=Sum('total'))['total'] or 0
            
            # Crear tabla de resúmenes
            total_general = advances_efectivo + advances_yape + advances_deposito + payments_efectivo_total + payments_yape_total
            
            summary_data = [
                ['CONCEPTO', 'MONTO'],
                ['INGRESOS DEL DÍA:', f"S/ {Decimal(total_advances):.2f}"],
                ['SALDOS:', f"S/ {Decimal(total_payments):.2f}"],
                ['SUBTOTAL INGRESOS:', f"S/ {float(total_advances + total_payments):.2f}"],
                ['TOTAL EGRESOS:', f"S/ {Decimal(total_expenses_amount):.2f}"],
                ['TOTAL EFECTIVO:', f"S/ {Decimal(advances_efectivo + payments_efectivo_total):.2f}"],
                ['TOTAL YAPE:', f"S/ {Decimal(advances_yape + payments_yape_total):.2f}"],
                ['TOTAL FINAL:', f"S/ {Decimal(total_general - total_expenses_amount):.2f}"]
            ]
            
            # Crear tabla de resúmenes
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.purple),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),  # Hacer toda la tabla en negrita
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # Construir PDF
            doc.build(story)
            
            # URL del archivo
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte PDF generado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al generar el PDF: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)